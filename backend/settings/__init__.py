"""
Settings, Payment Stages, Notes, Overdue, Activity Logs, Projects, and Units routes for RRL CRM.
"""
from typing import Optional, Dict, Any, List
from datetime import datetime, timezone
from fastapi import APIRouter, HTTPException, Depends
from pydantic import BaseModel, Field, ConfigDict
import uuid
import logging

from database import get_database
from config import settings as app_settings
from utils.enums import UserRole
from utils.payment_helpers import PAYMENT_STAGES
from auth import get_current_user, log_activity, check_role

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Settings & Stages"])
notes_router = APIRouter(tags=["Customer Notes"])
overdue_router = APIRouter(tags=["Overdue"])
misc_router = APIRouter(tags=["Misc"])
followups_router = APIRouter(tags=["Follow-ups"])

# Valid call-status values for the multi-level follow-up tracker.
# Tied to disbursement/payment stages from PAYMENT_STAGES.
FOLLOW_UP_STATUSES = ["Dialed", "Connected", "Unanswered", "Follow-up", "Completed"]


class UnitPricing(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project: str
    tower: str
    unit_number: str
    floor: int
    bhk_type: str
    saleable_area: float
    rate_per_sqft: float
    uds: float = 0
    is_available: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))


class UnitPricingCreate(BaseModel):
    project: str
    tower: str
    unit_number: str
    floor: int
    bhk_type: str
    saleable_area: float
    rate_per_sqft: float


# ==================== PAYMENT STAGE MANAGEMENT ====================
@router.get("/settings/payment-stages")
async def get_payment_stages(user: dict = Depends(get_current_user)):
    return PAYMENT_STAGES


@router.get("/settings/current-stage")
async def get_current_stage(user: dict = Depends(get_current_user)):
    db = get_database()
    settings_doc = await db.settings.find_one({"type": "payment_stage"}, {"_id": 0})
    if not settings_doc:
        return {"current_stage": None, "current_stage_name": None, "cumulative_percentage": 0}
    stage_key = settings_doc.get("current_stage")
    stage_info = next((s for s in PAYMENT_STAGES if s["key"] == stage_key), None)
    return {
        "current_stage": stage_key,
        "current_stage_name": stage_info["name"] if stage_info else None,
        "cumulative_percentage": stage_info["cumulative"] if stage_info else 0,
        "updated_at": settings_doc.get("updated_at"),
        "updated_by": settings_doc.get("updated_by_name")
    }


@router.post("/settings/current-stage")
async def set_current_stage(data: dict, user: dict = Depends(check_role([UserRole.ADMIN]))):
    db = get_database()
    stage_key = data.get("current_stage")
    if not stage_key:
        raise HTTPException(status_code=400, detail="current_stage is required")
    stage_info = next((s for s in PAYMENT_STAGES if s["key"] == stage_key), None)
    if not stage_info:
        raise HTTPException(status_code=400, detail="Invalid stage key")
    await db.settings.update_one(
        {"type": "payment_stage"},
        {"$set": {
            "type": "payment_stage", "current_stage": stage_key,
            "updated_at": datetime.now(timezone.utc),
            "updated_by": user["id"], "updated_by_name": user.get("name", "Admin")
        }},
        upsert=True
    )
    await log_activity(user['id'], user['name'], "update", "settings", "payment_stage", f"Set current payment stage to: {stage_info['name']}")
    return {"message": "Payment stage updated", "current_stage": stage_key, "stage_name": stage_info["name"]}


# ==================== CUSTOMER NOTES ====================
@notes_router.get("/customers/{customer_id}/notes")
async def get_customer_notes(customer_id: str, user: dict = Depends(get_current_user)):
    db = get_database()
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0, "notes": 1})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return customer.get("notes", [])


@notes_router.post("/customers/{customer_id}/notes")
async def add_customer_note(customer_id: str, data: dict, user: dict = Depends(get_current_user)):
    db = get_database()
    content = data.get("content", "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="Note content is required")
    note = {
        "id": str(uuid.uuid4()), "content": content,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"], "created_by_name": user.get("name", "Unknown")
    }
    result = await db.customers.update_one({"id": customer_id}, {"$push": {"notes": note}})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    await log_activity(user['id'], user['name'], "create", "note", customer_id, "Added note to customer")
    return note


@notes_router.delete("/customers/{customer_id}/notes/{note_id}")
async def delete_customer_note(customer_id: str, note_id: str, user: dict = Depends(get_current_user)):
    db = get_database()
    result = await db.customers.update_one({"id": customer_id}, {"$pull": {"notes": {"id": note_id}}})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Note not found")
    await log_activity(user['id'], user['name'], "delete", "note", customer_id, "Deleted note from customer")
    return {"message": "Note deleted"}


@notes_router.put("/customers/{customer_id}/payment-due-date")
async def update_payment_due_date(customer_id: str, data: dict, user: dict = Depends(get_current_user)):
    db = get_database()
    due_date = data.get("payment_due_date")
    result = await db.customers.update_one(
        {"id": customer_id},
        {"$set": {"payment_due_date": due_date, "updated_at": datetime.now(timezone.utc)}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    await log_activity(user['id'], user['name'], "update", "payment_due_date", customer_id, f"Updated payment due date to {due_date}")
    return {"message": "Payment due date updated", "payment_due_date": due_date}


# ==================== MULTI-LEVEL FOLLOW-UP TRACKER ====================
async def _compute_overdue_stages(customer: dict) -> List[Dict[str, Any]]:
    """Return PAYMENT_STAGES entries where cumulative expected > total_received.
    Capped at the admin-set current stage (anything beyond that is "not yet due").
    """
    db = get_database()
    settings_doc = await db.settings.find_one({"type": "payment_stage"}, {"_id": 0})
    current_stage_key = settings_doc.get("current_stage") if settings_doc else None
    if not current_stage_key:
        return []

    txns = await db.payment_transactions.find(
        {"customer_id": customer["id"]}, {"_id": 0, "amount": 1}
    ).to_list(1000)
    total_received = sum(t.get("amount", 0) or 0 for t in txns)
    total_price = customer.get("total_price", 0) or 0

    overdue = []
    for stage in PAYMENT_STAGES:
        expected = (total_price * stage["cumulative"]) / 100
        if expected > total_received + 1:  # tolerance
            overdue.append({**stage, "expected": round(expected, 2),
                            "shortfall": round(expected - total_received, 2)})
        if stage["key"] == current_stage_key:
            break
    return overdue


@followups_router.get("/customers/{customer_id}/follow-ups")
async def get_customer_follow_ups(customer_id: str, user: dict = Depends(get_current_user)):
    """Return follow-up log + overdue stages snapshot for the customer."""
    db = get_database()
    customer = await db.customers.find_one(
        {"id": customer_id},
        {"_id": 0, "id": 1, "name": 1, "follow_ups": 1, "total_price": 1}
    )
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")

    follow_ups = customer.get("follow_ups", []) or []
    overdue_stages = await _compute_overdue_stages(customer)

    settings_doc = await db.settings.find_one({"type": "payment_stage"}, {"_id": 0})
    current_stage_key = settings_doc.get("current_stage") if settings_doc else None

    return {
        "follow_ups": sorted(follow_ups, key=lambda f: f.get("created_at", ""), reverse=True),
        "overdue_stages": overdue_stages,
        "current_stage": current_stage_key,
        "all_stages": PAYMENT_STAGES,
        "statuses": FOLLOW_UP_STATUSES,
    }


@followups_router.post("/customers/{customer_id}/follow-ups")
async def add_customer_follow_up(customer_id: str, data: dict, user: dict = Depends(get_current_user)):
    db = get_database()
    stage_key = (data.get("stage_key") or "").strip()
    status = (data.get("status") or "").strip()
    notes_text = (data.get("notes") or "").strip()
    next_date = (data.get("next_follow_up_date") or "").strip() or None
    next_time = (data.get("next_follow_up_time") or "").strip() or None

    stage_info = next((s for s in PAYMENT_STAGES if s["key"] == stage_key), None)
    if not stage_info:
        raise HTTPException(status_code=400, detail="Invalid stage_key")
    if status not in FOLLOW_UP_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of {FOLLOW_UP_STATUSES}")

    entry = {
        "id": str(uuid.uuid4()),
        "stage_key": stage_key,
        "stage_name": stage_info["name"],
        "status": status,
        "notes": notes_text,
        "next_follow_up_date": next_date,
        "next_follow_up_time": next_time,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"],
        "created_by_name": user.get("name", "Unknown"),
    }
    result = await db.customers.update_one(
        {"id": customer_id}, {"$push": {"follow_ups": entry}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    await log_activity(
        user['id'], user['name'], "create", "follow_up", customer_id,
        f"Logged follow-up [{stage_info['name']}] status={status}"
    )
    return entry


@followups_router.delete("/customers/{customer_id}/follow-ups/{follow_up_id}")
async def delete_customer_follow_up(customer_id: str, follow_up_id: str, user: dict = Depends(get_current_user)):
    db = get_database()
    result = await db.customers.update_one(
        {"id": customer_id}, {"$pull": {"follow_ups": {"id": follow_up_id}}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Follow-up not found")
    await log_activity(user['id'], user['name'], "delete", "follow_up", customer_id, "Deleted follow-up")
    return {"message": "Follow-up deleted"}


@followups_router.post("/customers/{customer_id}/follow-ups/quick-status")
async def quick_set_call_status(customer_id: str, data: dict, user: dict = Depends(get_current_user)):
    """One-click call-status update from the Customers list view.

    Pushes a new follow-up entry against the admin-set ``current_stage`` so
    the column stays tied to disbursement progression. If admin hasn't set a
    current stage yet we fall back to the first PAYMENT_STAGE.
    """
    db = get_database()
    status = (data.get("status") or "").strip()
    notes_text = (data.get("notes") or "").strip()
    if status not in FOLLOW_UP_STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of {FOLLOW_UP_STATUSES}")

    settings_doc = await db.settings.find_one({"type": "payment_stage"}, {"_id": 0})
    stage_key = (settings_doc or {}).get("current_stage") or PAYMENT_STAGES[0]["key"]
    stage_info = next((s for s in PAYMENT_STAGES if s["key"] == stage_key), PAYMENT_STAGES[0])

    entry = {
        "id": str(uuid.uuid4()),
        "stage_key": stage_info["key"],
        "stage_name": stage_info["name"],
        "status": status,
        "notes": notes_text,
        "next_follow_up_date": None,
        "next_follow_up_time": None,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"],
        "created_by_name": user.get("name", "Unknown"),
    }
    result = await db.customers.update_one(
        {"id": customer_id}, {"$push": {"follow_ups": entry}}
    )
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    await log_activity(
        user['id'], user['name'], "create", "follow_up", customer_id,
        f"Quick-set call status [{stage_info['name']}] → {status}"
    )
    return entry


@followups_router.get("/follow-ups/upcoming")
async def get_upcoming_follow_ups(within_minutes: int = 120, user: dict = Depends(get_current_user)):
    """Return follow-ups whose ``next_follow_up_date[/time]`` falls within the
    next ``within_minutes`` window (or is already past-due today). Used by the
    in-app reminder hook to play a sound + browser notification.
    """
    db = get_database()
    now = datetime.now(timezone.utc)
    today_str = now.date().isoformat()
    cursor = db.customers.find(
        {"follow_ups": {"$exists": True, "$ne": []}},
        {"_id": 0, "id": 1, "name": 1, "customer_id": 1, "follow_ups": 1}
    )
    upcoming = []
    async for customer in cursor:
        for fu in customer.get("follow_ups", []) or []:
            d = fu.get("next_follow_up_date")
            if not d:
                continue
            if d > today_str:
                continue
            # Today or past-due → surface it
            upcoming.append({
                "follow_up_id": fu.get("id"),
                "customer_id": customer.get("id"),
                "customer_name": customer.get("name"),
                "stage_name": fu.get("stage_name"),
                "status": fu.get("status"),
                "notes": fu.get("notes"),
                "next_follow_up_date": d,
                "next_follow_up_time": fu.get("next_follow_up_time"),
                "is_today": d == today_str,
                "is_past_due": d < today_str,
            })
    upcoming.sort(key=lambda x: (x["next_follow_up_date"], x.get("next_follow_up_time") or "23:59"))
    return upcoming


# ==================== OVERDUE CALCULATION ====================
@overdue_router.get("/customers/{customer_id}/overdue")
async def get_customer_overdue(customer_id: str, user: dict = Depends(get_current_user)):
    db = get_database()
    settings_doc = await db.settings.find_one({"type": "payment_stage"}, {"_id": 0})
    if not settings_doc or not settings_doc.get("current_stage"):
        return {"overdue_amount": 0, "current_stage": None, "message": "No payment stage set by admin"}
    stage_key = settings_doc.get("current_stage")
    stage_info = next((s for s in PAYMENT_STAGES if s["key"] == stage_key), None)
    if not stage_info:
        return {"overdue_amount": 0, "current_stage": stage_key}
    cumulative_percentage = stage_info["cumulative"]
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    total_price = customer.get("total_price", 0) or 0
    expected_amount = (total_price * cumulative_percentage) / 100
    transactions = await db.payment_transactions.find({"customer_id": customer_id}, {"_id": 0}).to_list(1000)
    total_received = sum(t.get("amount", 0) or 0 for t in transactions)
    overdue_amount = max(0, expected_amount - total_received)
    return {
        "customer_id": customer_id, "current_stage": stage_key,
        "current_stage_name": stage_info["name"], "cumulative_percentage": cumulative_percentage,
        "expected_amount": round(expected_amount, 2), "total_received": round(total_received, 2),
        "overdue_amount": round(overdue_amount, 2), "is_overdue": overdue_amount > 0
    }


@overdue_router.get("/customers/overdue/list")
async def get_overdue_customers_list(user: dict = Depends(get_current_user)):
    overdue_data = await _get_overdue_by_stage_data(user)
    return {"customer_ids": [c["customer_id"] for c in overdue_data.get("overdue_customers", [])]}


@overdue_router.get("/dashboard/overdue-by-stage")
async def get_overdue_by_stage(user: dict = Depends(get_current_user)):
    return await _get_overdue_by_stage_data(user)


async def _get_overdue_by_stage_data(user: dict):
    db = get_database()
    settings_doc = await db.settings.find_one({"type": "payment_stage"}, {"_id": 0})
    if not settings_doc or not settings_doc.get("current_stage"):
        return {
            "current_stage": None, "overdue_count": 0, "total_overdue_amount": 0,
            "total_expected_at_slab": 0, "total_collected_cumulative": 0,
            "overdue_customers": []
        }
    stage_key = settings_doc.get("current_stage")
    stage_info = next((s for s in PAYMENT_STAGES if s["key"] == stage_key), None)
    if not stage_info:
        return {
            "current_stage": stage_key, "overdue_count": 0, "total_overdue_amount": 0,
            "total_expected_at_slab": 0, "total_collected_cumulative": 0,
            "overdue_customers": []
        }
    cumulative_percentage = stage_info["cumulative"]
    customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
    all_transactions = await db.payment_transactions.find({}, {"_id": 0}).to_list(100000)
    txn_by_customer = {}
    for txn in all_transactions:
        cid = txn.get("customer_id")
        if cid not in txn_by_customer:
            txn_by_customer[cid] = []
        txn_by_customer[cid].append(txn)

    overdue_customers = []
    total_overdue_amount = 0
    total_expected_at_slab = 0
    total_collected_cumulative = 0
    for customer in customers:
        customer_id = customer.get("id")
        total_price = customer.get("total_price", 0) or 0
        expected_amount = (total_price * cumulative_percentage) / 100
        customer_txns = txn_by_customer.get(customer_id, [])
        total_received = sum(t.get("amount", 0) or 0 for t in customer_txns)
        total_expected_at_slab += expected_amount
        total_collected_cumulative += total_received
        overdue_amount = expected_amount - total_received
        if overdue_amount > 0:
            overdue_customers.append({
                "customer_id": customer_id, "customer_name": customer.get("name"),
                "project": customer.get("project"), "unit_number": customer.get("unit_number"),
                "tower": customer.get("tower"), "total_price": total_price,
                "expected_amount": round(expected_amount, 2), "total_received": round(total_received, 2),
                "overdue_amount": round(overdue_amount, 2),
                "phone": customer.get("phone"), "email": customer.get("email"),
            })
            total_overdue_amount += overdue_amount
    overdue_customers.sort(key=lambda x: x["overdue_amount"], reverse=True)
    return {
        "current_stage": stage_key, "current_stage_name": stage_info["name"],
        "cumulative_percentage": cumulative_percentage,
        "overdue_count": len(overdue_customers),
        "total_overdue_amount": round(total_overdue_amount, 2),
        "total_expected_at_slab": round(total_expected_at_slab, 2),
        "total_collected_cumulative": round(total_collected_cumulative, 2),
        "overdue_customers": overdue_customers
    }


# ==================== ACTIVITY LOGS ====================
@misc_router.get("/activity-logs")
async def get_activity_logs(
    entity_type: Optional[str] = None, entity_id: Optional[str] = None,
    skip: int = 0, limit: int = 100, user: dict = Depends(get_current_user)
):
    db = get_database()
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    if entity_id:
        query["entity_id"] = entity_id
    logs = await db.activity_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)
    return logs


# ==================== PROJECTS ====================
@misc_router.get("/projects")
async def get_projects(user: dict = Depends(get_current_user)):
    return [
        {"name": "RRL Palm Altezze", "location": "Varthur, Bangalore"},
        {"name": "RRL NC 216", "location": "Bangalore"},
        {"name": "RRL Palacio", "location": "Medahalli, Bangalore"},
        {"name": "RRL Nature Woods", "location": "Sarjapur, Bangalore"},
        {"name": "RRL Towers", "location": "Sarjapur"},
        {"name": "RRL Complex", "location": "Attibele Sarjapur Road"}
    ]


# ==================== UNIT PRICING ====================
@misc_router.post("/units")
async def create_unit_pricing(unit: UnitPricingCreate, user: dict = Depends(check_role([UserRole.ADMIN, UserRole.MANAGER]))):
    db = get_database()
    unit_doc = UnitPricing(**unit.model_dump(), uds=round(unit.saleable_area * 0.495046, 2))
    doc = unit_doc.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    await db.units.insert_one(doc)
    await log_activity(user['id'], user['name'], "create", "unit", unit_doc.id, f"Created unit {unit.unit_number}")
    return {**doc, "_id": None}


@misc_router.get("/units")
async def get_units(
    project: Optional[str] = None, tower: Optional[str] = None,
    bhk_type: Optional[str] = None, is_available: Optional[bool] = None,
    user: dict = Depends(get_current_user)
):
    db = get_database()
    query = {}
    if project:
        query["project"] = project
    if tower:
        query["tower"] = tower
    if bhk_type:
        query["bhk_type"] = bhk_type
    if is_available is not None:
        query["is_available"] = is_available
    units = await db.units.find(query, {"_id": 0}).to_list(1000)
    return units


@misc_router.get("/units/{unit_id}")
async def get_unit(unit_id: str, user: dict = Depends(get_current_user)):
    db = get_database()
    unit = await db.units.find_one({"id": unit_id}, {"_id": 0})
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")
    return unit


@misc_router.put("/units/{unit_id}")
async def update_unit(unit_id: str, updates: Dict[str, Any], user: dict = Depends(check_role([UserRole.ADMIN, UserRole.MANAGER]))):
    db = get_database()
    if 'saleable_area' in updates:
        updates['uds'] = round(updates['saleable_area'] * 0.495046, 2)
    result = await db.units.update_one({"id": unit_id}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Unit not found")
    await log_activity(user['id'], user['name'], "update", "unit", unit_id, "Updated unit")
    return {"message": "Unit updated"}


@misc_router.post("/units/bulk-import")
async def bulk_import_units(units: List[UnitPricingCreate], user: dict = Depends(check_role([UserRole.ADMIN]))):
    db = get_database()
    created = 0
    for unit_data in units:
        unit_doc = UnitPricing(**unit_data.model_dump(), uds=round(unit_data.saleable_area * 0.495046, 2))
        doc = unit_doc.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.units.insert_one(doc)
        created += 1
    await log_activity(user['id'], user['name'], "import", "units", "", f"Imported {created} units")
    return {"message": f"Imported {created} units successfully"}


# ==================== EXPORT FUNCTIONALITY ====================
@misc_router.get("/export/customers/csv")
async def export_customers_csv(user: dict = Depends(check_role([UserRole.ADMIN, UserRole.MANAGER]))):
    import io
    import csv
    from fastapi.responses import Response

    db = get_database()
    customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
    if not customers:
        raise HTTPException(status_code=404, detail="No customers found")

    output = io.StringIO()
    headers = [
        "Customer ID", "Name", "Email", "Phone", "Project", "Tower", "Unit Number",
        "BHK Type", "Floor", "Saleable Area", "Rate/Sqft", "Base Price", "Floor Rise Cost",
        "Club House", "Additional Parking", "Labour Cess", "GST Amount", "Total Price",
        "Booking Amount", "Booking Date", "Total Received", "Balance Amount",
        "Payment Received %", "Agreement Status", "Stage", "Father Name", "PAN Number",
        "Address", "Created At"
    ]
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    for c in customers:
        writer.writerow({
            "Customer ID": c.get('customer_id', ''), "Name": c.get('name', ''),
            "Email": c.get('email', ''), "Phone": c.get('phone', ''),
            "Project": c.get('project', ''), "Tower": c.get('tower', ''),
            "Unit Number": c.get('unit_number', ''), "BHK Type": c.get('unit_type', ''),
            "Floor": c.get('floor', ''), "Saleable Area": c.get('saleable_area', 0),
            "Rate/Sqft": c.get('rate_per_sqft', 0), "Base Price": c.get('base_price', 0),
            "Floor Rise Cost": c.get('custom_fields', {}).get('floor_rise_cost', 0),
            "Club House": c.get('club_house_charges', 0),
            "Additional Parking": c.get('additional_parking_charges', 0),
            "Labour Cess": c.get('labour_cess', 0), "GST Amount": c.get('gst_amount', 0),
            "Total Price": c.get('total_price', 0), "Booking Amount": c.get('booking_amount', 0),
            "Booking Date": c.get('booking_date', ''), "Total Received": c.get('total_received', 0),
            "Balance Amount": c.get('balance_amount', 0),
            "Payment Received %": c.get('payment_received_percentage', 0),
            "Agreement Status": c.get('agreement_status', ''), "Stage": c.get('stage', ''),
            "Father Name": c.get('father_name', ''), "PAN Number": c.get('pan_number', ''),
            "Address": c.get('address', ''), "Created At": c.get('created_at', '')
        })
    csv_content = output.getvalue()
    output.close()
    await log_activity(user['id'], user['name'], "export", "customers", "all", "Exported customers to CSV")
    return Response(content=csv_content, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=RRL_Customers_Export.csv"})


@misc_router.get("/export/customers/excel")
async def export_customers_excel(user: dict = Depends(check_role([UserRole.ADMIN, UserRole.MANAGER]))):
    import io
    from fastapi.responses import Response
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export not available. Please install openpyxl.")

    db = get_database()
    customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
    if not customers:
        raise HTTPException(status_code=404, detail="No customers found")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    headers = [
        "Customer ID", "Name", "Email", "Phone", "Project", "Tower", "Unit Number",
        "BHK Type", "Floor", "Saleable Area", "Rate/Sqft", "Base Price", "Floor Rise Cost",
        "Club House", "Additional Parking", "Labour Cess", "GST Amount", "Total Price",
        "Booking Amount", "Booking Date", "Total Received", "Balance Amount",
        "Payment Received %", "Agreement Status", "Stage", "Created At"
    ]
    header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    header_font = Font(color="D4AF37", bold=True, name="Roboto")
    thin_border = Border(
        left=Side(style='thin', color='D4AF37'), right=Side(style='thin', color='D4AF37'),
        top=Side(style='thin', color='D4AF37'), bottom=Side(style='thin', color='D4AF37')
    )
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    for row, c in enumerate(customers, 2):
        data = [
            c.get('customer_id', ''), c.get('name', ''), c.get('email', ''), c.get('phone', ''),
            c.get('project', ''), c.get('tower', ''), c.get('unit_number', ''),
            c.get('unit_type', ''), c.get('floor', ''), c.get('saleable_area', 0),
            c.get('rate_per_sqft', 0), c.get('base_price', 0),
            c.get('custom_fields', {}).get('floor_rise_cost', 0),
            c.get('club_house_charges', 0), c.get('additional_parking_charges', 0),
            c.get('labour_cess', 0), c.get('gst_amount', 0), c.get('total_price', 0),
            c.get('booking_amount', 0), c.get('booking_date', ''), c.get('total_received', 0),
            c.get('balance_amount', 0), c.get('payment_received_percentage', 0),
            c.get('agreement_status', ''), c.get('stage', ''), c.get('created_at', '')
        ]
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = thin_border
            cell.font = Font(name="Roboto")
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    await log_activity(user['id'], user['name'], "export", "customers", "all", "Exported customers to Excel")
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=RRL_Customers_Export.xlsx"}
    )


@misc_router.get("/export/payments/csv")
async def export_payments_csv(user: dict = Depends(check_role([UserRole.ADMIN, UserRole.MANAGER, UserRole.ACCOUNTS]))):
    import io
    import csv
    from fastapi.responses import Response

    db = get_database()
    schedules = await db.payment_schedules.find({}, {"_id": 0}).to_list(10000)
    customers = {c['id']: c for c in await db.customers.find({}, {"_id": 0, "id": 1, "name": 1, "customer_id": 1, "project": 1, "unit_number": 1}).to_list(10000)}

    output = io.StringIO()
    headers = ["Customer ID", "Customer Name", "Project", "Unit", "Installment", "Milestone", "Amount", "Due Date", "Status", "Payment Date"]
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    for schedule in schedules:
        customer = customers.get(schedule.get('customer_id'), {})
        for item in schedule.get('items', []):
            writer.writerow({
                "Customer ID": customer.get('customer_id', ''), "Customer Name": customer.get('name', ''),
                "Project": customer.get('project', ''), "Unit": customer.get('unit_number', ''),
                "Installment": item.get('installment_name', ''), "Milestone": item.get('milestone', ''),
                "Amount": item.get('amount', 0), "Due Date": item.get('due_date', ''),
                "Status": item.get('payment_status', ''), "Payment Date": item.get('payment_date', '')
            })
    csv_content = output.getvalue()
    output.close()
    await log_activity(user['id'], user['name'], "export", "payments", "all", "Exported payments to CSV")
    return Response(content=csv_content, media_type="text/csv", headers={"Content-Disposition": "attachment; filename=RRL_Payments_Export.csv"})
