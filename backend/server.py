from fastapi import FastAPI, APIRouter, HTTPException, Depends, BackgroundTasks, UploadFile, File, Form
from fastapi import status as http_status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pymongo import ReturnDocument
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, ConfigDict
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
from enum import Enum
from io import BytesIO
from fastapi.responses import StreamingResponse, Response
import base64
import json

# SendGrid imports
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail, Attachment, FileContent, FileName, FileType, Disposition

# WeasyPrint for PDF generation
from weasyprint import HTML

# ==================== MODULAR IMPORTS ====================
# Importing from new modular structure
from config import settings
from database import get_database, db as module_db
from utils import number_to_indian_words as utils_number_to_indian_words
from utils import format_indian_currency as utils_format_indian_currency
from utils.enums import (
    UserRole as EnumUserRole, CustomerStage as EnumCustomerStage,
    AgreementStatus as EnumAgreementStatus, FinanceType as EnumFinanceType,
    PaymentStatus as EnumPaymentStatus, DocumentType as EnumDocumentType,
    TransactionStage as EnumTransactionStage
)

# Import modular routers
from auth import router as auth_router, admin_router as auth_admin_router, users_router
from auth import get_current_user as module_get_current_user, log_activity as module_log_activity
from auth import hash_password as module_hash_password, verify_password as module_verify_password
from auth import create_token as module_create_token, check_role as module_check_role
from customers import router as customers_router
from payments import schedule_router, transactions_router, calculator_router
from dashboard import router as dashboard_router

# Import document HTML template generators
from documents.templates import (
    generate_sales_agreement_template, get_default_template,
    generate_price_breakup_html, generate_cost_breakup_html,
    generate_noc_hdfc_html, generate_noc_bob_html, generate_noc_tata_html,
    generate_booking_form_preview_html, generate_terms_and_conditions_html,
    generate_welcome_email_html, generate_document_email_html,
    generate_sales_agreement_html, generate_allotment_letter_html,
    generate_payment_schedule_pdf_html, generate_payment_schedule_html,
    generate_demand_letter_html
)

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Settings (using config.settings)
JWT_SECRET = settings.JWT_SECRET
JWT_ALGORITHM = settings.JWT_ALGORITHM
JWT_EXPIRATION_HOURS = settings.JWT_EXPIRATION_HOURS

# SendGrid Settings (using config.settings)
SENDGRID_API_KEY = settings.SENDGRID_API_KEY
SENDGRID_FROM_EMAIL = settings.SENDGRID_FROM_EMAIL
SENDGRID_FROM_NAME = settings.SENDGRID_FROM_NAME

# Create the main app
app = FastAPI(title="RRL Builders CRM API")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Utility function to convert number to Indian words
def number_to_indian_words(num):
    """Convert number to words in Indian format (Lakhs, Crores)"""
    if num == 0:
        return "Zero"
    
    ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine",
            "Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen",
            "Seventeen", "Eighteen", "Nineteen"]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
    
    def convert_less_than_thousand(n):
        if n == 0:
            return ""
        elif n < 20:
            return ones[n]
        elif n < 100:
            return tens[n // 10] + (" " + ones[n % 10] if n % 10 != 0 else "")
        else:
            return ones[n // 100] + " Hundred" + (" " + convert_less_than_thousand(n % 100) if n % 100 != 0 else "")
    
    num = int(num)
    if num < 0:
        return "Minus " + number_to_indian_words(-num)
    
    crore = num // 10000000
    lakh = (num % 10000000) // 100000
    thousand = (num % 100000) // 1000
    remainder = num % 1000
    
    result = ""
    if crore > 0:
        result += convert_less_than_thousand(crore) + " Crore "
    if lakh > 0:
        result += convert_less_than_thousand(lakh) + " Lakh "
    if thousand > 0:
        result += convert_less_than_thousand(thousand) + " Thousand "
    if remainder > 0:
        result += convert_less_than_thousand(remainder)
    
    return result.strip() + " Rupees"

def format_indian_currency(amount, decimals=True):
    """Format amount in Indian currency style (e.g., 12,34,567.00 or 12,34,567)"""
    amount = float(amount) if amount else 0
    int_part = int(amount)
    
    # Format integer part with Indian comma system
    s = str(int_part)
    if len(s) > 3:
        # First group of 3, then groups of 2
        result = s[-3:]
        s = s[:-3]
        while s:
            result = s[-2:] + ',' + result
            s = s[:-2]
    else:
        result = s
    
    if decimals:
        decimal_part = f"{amount:.2f}".split('.')[1]
        return f"{result}.{decimal_part}"
    return result

# Security
security = HTTPBearer()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== ENUMS ====================
class UserRole(str, Enum):
    ADMIN = "admin"
    MANAGER = "manager"
    ACCOUNTS = "accounts"
    SALES = "sales"
    SUPPORT = "support"

class CustomerStage(str, Enum):
    PENDING_APPROVAL = "pending_approval"
    QUALIFIED = "qualified"
    AGREEMENT_PENDING = "agreement_pending"
    AGREEMENT_DONE = "agreement_done"
    REGISTRATION_DONE = "registration_done"

class AgreementStatus(str, Enum):
    DRAFT = "draft"
    SENT = "sent"
    SIGNED = "signed"
    COMPLETED = "completed"

class FinanceType(str, Enum):
    SELF = "self"
    LOAN = "loan"
    MIXED = "mixed"

class PaymentStatus(str, Enum):
    PENDING = "pending"
    PAID = "paid"
    OVERDUE = "overdue"
    PARTIAL = "partial"

class DocumentType(str, Enum):
    SALES_AGREEMENT = "sales_agreement"
    ALLOTMENT_LETTER = "allotment_letter"
    DISBURSEMENT_LETTER = "disbursement_letter"
    PRICE_BREAKUP = "price_breakup"
    COST_BREAKUP = "cost_breakup"
    WELCOME_LETTER = "welcome_letter"
    DEMAND_LETTER = "demand_letter"
    PAYMENT_SCHEDULE = "payment_schedule"
    # Bank NOC Documents for Disbursement
    NOC_HDFC = "noc_hdfc"
    NOC_BOB = "noc_bob"
    NOC_TATA = "noc_tata"

# ==================== UNIT PRICING MODEL ====================
class UnitPricing(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    project: str
    tower: str
    unit_number: str
    floor: int
    bhk_type: str  # 2BHK, 3BHK
    saleable_area: float
    rate_per_sqft: float
    uds: float = 0  # Undivided Share - calculated as saleable_area * 0.495046
    is_available: bool = True
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UnitPricingCreate(BaseModel):
    project: str
    tower: str
    unit_number: str
    floor: int
    bhk_type: str
    saleable_area: float
    rate_per_sqft: float

# ==================== MODELS ====================
class UserBase(BaseModel):
    email: EmailStr
    name: str
    role: UserRole = UserRole.SALES
    phone: Optional[str] = None
    is_active: bool = True

class UserCreate(UserBase):
    password: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class User(UserBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserResponse(BaseModel):
    id: str
    email: str
    name: str
    role: UserRole
    phone: Optional[str] = None
    is_active: bool
    created_at: str

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

class CustomerBase(BaseModel):
    # Primary Applicant
    name: str
    phone: str
    email: EmailStr
    father_name: Optional[str] = None
    date_of_birth: Optional[str] = None
    gender: Optional[str] = None  # male, female, or spouse (for W/o)
    pan_number: Optional[str] = None
    aadhar_number: Optional[str] = None
    address: Optional[str] = None
    company: Optional[str] = None
    designation: Optional[str] = None
    nationality: str = "Indian"
    
    # Co-Applicant Details
    co_applicant_name: Optional[str] = None
    co_applicant_father_name: Optional[str] = None
    co_applicant_phone: Optional[str] = None
    co_applicant_email: Optional[str] = None
    co_applicant_pan: Optional[str] = None
    co_applicant_aadhar: Optional[str] = None
    co_applicant_address: Optional[str] = None
    co_applicant_date_of_birth: Optional[str] = None
    
    # Property Details
    project: str
    tower: str
    unit_number: str
    booking_number: Optional[str] = None  # Unique booking ID (e.g., RRL PAB035)
    floor: int = 0
    bhk_type: str = ""
    saleable_area: float = 0
    uds: float = 0  # Undivided Share
    parking: Optional[str] = None
    additional_parking: int = 0  # Number of additional parking (legacy)
    
    # Pricing
    rate_per_sqft: float = 0
    base_price: float = 0  # rate * saleable_area
    club_house_charges: float = 200000  # Default 200000, editable
    infrastructure_charges: float = 0
    additional_charges: float = 0  # Manual additional charges
    additional_parking_charges: float = 0  # Legacy field
    labour_cess: float = 0  # 0.70%
    gst_percentage: float = 5
    gst_amount: float = 0
    total_price: float = 0  # Total including GST
    
    # Payment Tracking
    booking_amount: float = 0
    booking_date: Optional[str] = None
    agreement_date: Optional[str] = None
    total_received: float = 0
    balance_amount: float = 0
    payment_received_percentage: float = 0
    payment_pending_percentage: float = 100
    
    # Finance Details
    finance_type: str = "self"  # self, loan, mixed
    finance_bank: Optional[str] = None
    loan_amount: float = 0
    self_contribution: float = 0
    first_disbursement_amount: float = 0
    first_disbursement_date: Optional[str] = None
    
    # Bank Account Details (Customer's bank for disbursement/refund)
    bank_name: Optional[str] = None  # HDFC, BOB, TATA, Others
    bank_name_other: Optional[str] = None  # If bank_name is "Others"
    bank_account_number: Optional[str] = None
    bank_ifsc_code: Optional[str] = None
    bank_branch: Optional[str] = None
    bank_account_holder: Optional[str] = None
    
    # Status
    stage: str = "pending_approval"  # pending_approval, qualified, agreement_pending, agreement_done, registration_done
    agreement_status: str = "draft"
    ownership: str = "builder"  # builder, customer
    registration_date: Optional[str] = None
    handover_date: Optional[str] = None
    
    # Transaction Details
    transaction_details: Optional[str] = None
    transaction_date: Optional[str] = None
    transaction_bank: Optional[str] = None
    
    # Notes
    remarks: Optional[str] = None
    custom_fields: Dict[str, Any] = {}
    
    # Document Uploads (store file paths/urls)
    uploaded_documents: Dict[str, str] = {}  # {doc_type: file_url}

class CustomerCreate(CustomerBase):
    pass

class Customer(CustomerBase):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: Optional[str] = None

class PaymentScheduleItem(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    installment_name: str
    milestone: str
    amount: float
    due_date: str
    payment_status: PaymentStatus = PaymentStatus.PENDING
    payment_date: Optional[str] = None
    bank_disbursement_status: Optional[str] = None

class PaymentScheduleCreate(BaseModel):
    customer_id: str
    items: List[PaymentScheduleItem]

class PaymentSchedule(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str
    items: List[PaymentScheduleItem] = []
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DocumentTemplate(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    doc_type: DocumentType
    content: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DocumentGenerate(BaseModel):
    customer_id: str
    doc_type: DocumentType
    custom_fields: Dict[str, str] = {}

class GeneratedDocument(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str
    doc_type: DocumentType
    content: str
    generated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    generated_by: str
    signed_copy_url: Optional[str] = None
    status: AgreementStatus = AgreementStatus.DRAFT

class DocumentChecklist(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str
    items: Dict[str, bool] = {
        "kyc_documents": False,
        "pan_card": False,
        "aadhar": False,
        "agreement_copy": False,
        "bank_documents": False,
        "passport_photo": False,
        "address_proof": False
    }
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CommunicationLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str
    channel: str  # email, whatsapp
    message_type: str
    content: str
    status: str = "sent"
    sent_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    sent_by: str

class ActivityLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    user_name: str
    action: str
    entity_type: str
    entity_id: str
    details: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TransactionStage(str, Enum):
    BOOKING = "booking"
    AGREEMENT = "agreement"
    SCHEDULED_DISBURSEMENT = "scheduled_disbursement"

class PaymentTransaction(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    customer_id: str
    transaction_stage: TransactionStage
    transaction_date: str
    bank_name: str
    transaction_number: str
    amount: Optional[float] = 0
    notes: Optional[str] = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PaymentTransactionCreate(BaseModel):
    transaction_stage: TransactionStage
    transaction_date: str
    bank_name: str
    transaction_number: str
    amount: Optional[float] = 0
    notes: Optional[str] = ""

class PriceCalculation(BaseModel):
    """Enhanced price calculation matching the Excel formula"""
    unit_number: Optional[str] = None
    unit_type: Optional[str] = None  # 2BHK, 3BHK
    floor_number: int = 0
    saleable_area: float  # sq.ft
    rate_per_sqft: float
    include_club_house: bool = True  # Rs. 2,00,000 if true
    club_house_charges: float = 200000  # Editable
    additional_charges: float = 0  # Manual additional charges
    additional_parking_count: int = 0  # Legacy field
    additional_parking_rate: float = 300000
    gst_percentage: float = 5
    labour_cess_percentage: float = 0.70

class PriceResult(BaseModel):
    # Input echo
    unit_number: Optional[str] = None
    unit_type: Optional[str] = None
    floor_number: int = 0
    saleable_area: float = 0
    rate_per_sqft: float = 0
    
    # Calculations
    base_price: float  # rate * saleable_area
    club_house_charges: float
    additional_charges: float = 0  # Manual additional charges
    additional_parking_charges: float = 0  # Legacy
    subtotal_before_taxes: float  # base + club + charges
    labour_cess: float  # 0.70% of subtotal
    gst_amount: float  # 5% of subtotal
    total_flat_value: float  # subtotal + taxes
    
    # UDS calculation
    uds: float  # saleable_area * 0.495046

class DisbursementCalculation(BaseModel):
    """For calculating disbursement amounts"""
    total_flat_value: float
    disbursement_percentage: float = 30  # Default 30%

class DisbursementResult(BaseModel):
    total_flat_value: float
    disbursement_percentage: float
    disbursement_amount: float

class PaymentTrackingResult(BaseModel):
    """Calculate payment tracking metrics"""
    total_flat_value: float
    total_received: float
    balance_amount: float
    payment_received_percentage: float
    payment_pending_percentage: float

class PaymentScheduleTemplate(BaseModel):
    """Payment schedule template from Excel"""
    installment_name: str
    percentage: float
    milestone: str

class GoogleFormWebhook(BaseModel):
    customer_name: str
    phone: str
    email: EmailStr
    project: str
    tower: str
    unit_number: str
    father_name: Optional[str] = None
    pan_number: Optional[str] = None
    booking_amount: Optional[float] = 0
    booking_date: Optional[str] = None

class DashboardStats(BaseModel):
    total_customers: int
    pending_agreements: int
    payments_due_this_week: int
    overdue_payments: int
    total_revenue: float
    total_pending: float
    total_flat_value: float
    total_balance: float
    pending_percentage: float
    monthly_revenue: List[Dict[str, Any]]
    payment_status_breakdown: Dict[str, int]
    # New fields for stage-based overdue tracking
    current_stage: Optional[str] = None
    current_stage_name: Optional[str] = None
    stage_overdue_count: int = 0
    stage_overdue_amount: float = 0
    overdue_customers: List[Dict[str, Any]] = []

# ==================== PAYMENT STAGE SETTINGS ====================
class PaymentStageSettings(BaseModel):
    current_stage: str  # milestone key like "podium", "2nd_floor", etc.
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_by: Optional[str] = None

class CustomerNote(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    content: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    created_by: str
    created_by_name: str

# Payment stages available for admin selection (from Podium Slab to Possession)
PAYMENT_STAGES = [
    {"key": "podium", "name": "On Completion of Podium Slab", "percentage": 40, "cumulative": 40},
    {"key": "2nd_floor", "name": "Upon Completion of 2nd Floor Roof Slab", "percentage": 5, "cumulative": 45},
    {"key": "6th_floor", "name": "Upon Completion of 6th Floor Roof Slab", "percentage": 5, "cumulative": 50},
    {"key": "10th_floor", "name": "Upon Completion of 10th Floor Roof Slab", "percentage": 5, "cumulative": 55},
    {"key": "14th_floor", "name": "Upon Completion of 14th Floor Roof Slab", "percentage": 5, "cumulative": 60},
    {"key": "18th_floor", "name": "Upon Completion of 18th Floor Roof Slab", "percentage": 5, "cumulative": 65},
    {"key": "22nd_floor", "name": "Upon Completion of 22nd Floor Roof Slab", "percentage": 5, "cumulative": 70},
    {"key": "top_roof", "name": "Upon Completion of Top Roof Slab", "percentage": 10, "cumulative": 80},
    {"key": "flooring", "name": "Upon Completion of Flooring of Particular Property", "percentage": 10, "cumulative": 90},
    {"key": "handover", "name": "Upon Handover / Possession / Registration", "percentage": 10, "cumulative": 100},
]

# ==================== HELPER FUNCTIONS ====================
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def check_role(required_roles: List[UserRole]):
    async def role_checker(user: dict = Depends(get_current_user)):
        if UserRole(user["role"]) not in required_roles:
            raise HTTPException(status_code=403, detail="Insufficient permissions")
        return user
    return role_checker

async def log_activity(user_id: str, user_name: str, action: str, entity_type: str, entity_id: str, details: str):
    log = ActivityLog(
        user_id=user_id,
        user_name=user_name,
        action=action,
        entity_type=entity_type,
        entity_id=entity_id,
        details=details
    )
    doc = log.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    await db.activity_logs.insert_one(doc)

async def generate_customer_id():
    """Generate unique customer ID using atomic counter"""
    # Use findOneAndUpdate with upsert for atomic counter increment
    result = await db.counters.find_one_and_update(
        {"_id": "customer_id"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=ReturnDocument.AFTER
    )
    return f"RRL-{str(result['seq']).zfill(5)}"

# ==================== AUTH ROUTES ====================
@api_router.post("/auth/register", response_model=UserResponse)
async def register_user(user_data: UserCreate):
    existing = await db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    user = User(**user_data.model_dump(exclude={"password"}))
    doc = user.model_dump()
    doc['password_hash'] = hash_password(user_data.password)
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.users.insert_one(doc)
    return UserResponse(
        id=user.id,
        email=user.email,
        name=user.name,
        role=user.role,
        phone=user.phone,
        is_active=user.is_active,
        created_at=doc['created_at']
    )

@api_router.post("/auth/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user.get('password_hash', '')):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if not user.get('is_active', True):
        raise HTTPException(status_code=401, detail="Account disabled")
    
    token = create_token(user['id'], user['email'], user['role'])
    return TokenResponse(
        access_token=token,
        user=UserResponse(
            id=user['id'],
            email=user['email'],
            name=user['name'],
            role=user['role'],
            phone=user.get('phone'),
            is_active=user.get('is_active', True),
            created_at=user['created_at']
        )
    )

@api_router.get("/auth/me", response_model=UserResponse)
async def get_me(user: dict = Depends(get_current_user)):
    return UserResponse(
        id=user['id'],
        email=user['email'],
        name=user['name'],
        role=user['role'],
        phone=user.get('phone'),
        is_active=user.get('is_active', True),
        created_at=user['created_at']
    )

# Password Reset Models
class VerifyEmailRequest(BaseModel):
    email: EmailStr

class ResetPasswordRequest(BaseModel):
    email: EmailStr
    new_password: str

@api_router.post("/auth/verify-email")
async def verify_email(data: VerifyEmailRequest):
    """Verify if email exists in the system for password reset"""
    user = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Email not found in our system")
    return {"exists": True, "message": "Email verified"}

@api_router.post("/auth/reset-password")
async def reset_password(data: ResetPasswordRequest):
    """Reset user password"""
    user = await db.users.find_one({"email": data.email}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="Email not found in our system")
    
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    
    new_hash = hash_password(data.new_password)
    result = await db.users.update_one(
        {"email": data.email},
        {"$set": {"password_hash": new_hash}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=500, detail="Failed to update password")
    
    return {"message": "Password reset successfully"}

# Admin Reset User Password
class AdminResetPasswordRequest(BaseModel):
    user_id: str
    new_password: str

@api_router.post("/admin/reset-user-password")
async def admin_reset_user_password(data: AdminResetPasswordRequest, current_user: dict = Depends(get_current_user)):
    """Admin can reset any user's password"""
    if current_user['role'] != 'admin':
        raise HTTPException(status_code=403, detail="Only admins can reset user passwords")
    
    user = await db.users.find_one({"id": data.user_id}, {"_id": 0})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    if len(data.new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")
    
    new_hash = hash_password(data.new_password)
    result = await db.users.update_one(
        {"id": data.user_id},
        {"$set": {"password_hash": new_hash}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=500, detail="Failed to update password")
    
    return {"message": f"Password reset successfully for {user['name']}"}

# ==================== USER MANAGEMENT ROUTES ====================
@api_router.get("/users", response_model=List[UserResponse])
async def get_users(user: dict = Depends(check_role([UserRole.ADMIN]))):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return [UserResponse(**u) for u in users]

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, updates: Dict[str, Any], user: dict = Depends(check_role([UserRole.ADMIN]))):
    if 'password' in updates:
        updates['password_hash'] = hash_password(updates.pop('password'))
    
    result = await db.users.update_one({"id": user_id}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    await log_activity(user['id'], user['name'], "update", "user", user_id, "Updated user")
    return {"message": "User updated"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, user: dict = Depends(check_role([UserRole.ADMIN]))):
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    await log_activity(user['id'], user['name'], "delete", "user", user_id, "Deleted user")
    return {"message": "User deleted"}

# ==================== CUSTOMER ROUTES ====================
@api_router.post("/customers", response_model=Dict[str, Any])
async def create_customer(customer_data: CustomerCreate, user: dict = Depends(get_current_user)):
    customer = Customer(**customer_data.model_dump())
    customer.customer_id = await generate_customer_id()
    customer.created_by = user['id']
    
    doc = customer.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await db.customers.insert_one(doc)
    
    # Create default document checklist
    checklist = DocumentChecklist(customer_id=customer.id)
    checklist_doc = checklist.model_dump()
    checklist_doc['updated_at'] = checklist_doc['updated_at'].isoformat()
    await db.document_checklists.insert_one(checklist_doc)
    
    await log_activity(user['id'], user['name'], "create", "customer", customer.id, f"Created customer {customer.name}")
    
    # Auto-generate booking transaction if booking_amount is set
    await auto_generate_booking_transaction(customer.id, doc, created_by=user['id'])
    
    return {**doc, "_id": None}

@api_router.get("/customers")
async def get_customers(
    search: Optional[str] = None,
    project: Optional[str] = None,
    agreement_status: Optional[str] = None,
    agreement_filter: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    user: dict = Depends(get_current_user)
):
    query = {}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"email": {"$regex": search, "$options": "i"}},
            {"phone": {"$regex": search, "$options": "i"}},
            {"customer_id": {"$regex": search, "$options": "i"}},
            {"booking_number": {"$regex": search, "$options": "i"}},
            {"unit_number": {"$regex": search, "$options": "i"}}
        ]
    if project:
        query["project"] = project
    if agreement_status:
        query["agreement_status"] = agreement_status
    
    # Apply agreement filters
    filter_overdue = False
    if agreement_filter:
        today = datetime.now(timezone.utc).date()
        
        if agreement_filter == "upcoming_due":
            # Customers with due date in next 5 days (10 days from booking)
            # We'll filter in Python since we need date calculation
            pass
        elif agreement_filter == "pending_agreement":
            # Customers with draft or sent agreement status
            query["agreement_status"] = {"$in": ["draft", "sent"]}
        elif agreement_filter == "agreement_due":
            # Customers whose agreement needs signing (sent but not signed)
            query["agreement_status"] = "sent"
        elif agreement_filter == "overdue":
            # Customers who are overdue based on current stage
            filter_overdue = True
    
    customers = await db.customers.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit * 2 if agreement_filter in ["upcoming_due", "overdue"] else limit)
    
    # Post-filter for upcoming_due
    if agreement_filter == "upcoming_due":
        today = datetime.now(timezone.utc).date()
        filtered_customers = []
        
        for customer in customers:
            booking_date_str = customer.get('booking_date')
            if not booking_date_str:
                continue
            try:
                if isinstance(booking_date_str, str):
                    booking_date = datetime.strptime(booking_date_str, "%Y-%m-%d").date()
                else:
                    booking_date = booking_date_str
                
                # Due date is 10 days from booking
                due_date = booking_date + timedelta(days=10)
                days_until_due = (due_date - today).days
                
                # Include if due within next 5 days (including recently overdue up to 3 days)
                if -3 <= days_until_due <= 5:
                    customer['_due_date'] = due_date.isoformat()
                    customer['_days_until_due'] = days_until_due
                    filtered_customers.append(customer)
            except Exception:
                continue
        
        # Sort by due date (closest first)
        filtered_customers.sort(key=lambda x: x.get('_days_until_due', 999))
        customers = filtered_customers[:limit]
    
    # Post-filter for overdue customers based on current stage
    if filter_overdue:
        settings = await db.settings.find_one({"type": "payment_stage"}, {"_id": 0})
        if settings and settings.get("current_stage"):
            stage_key = settings.get("current_stage")
            stage_info = next((s for s in PAYMENT_STAGES if s["key"] == stage_key), None)
            if stage_info:
                cumulative_percentage = stage_info["cumulative"]
                
                # Get transactions only for current page customers
                page_cust_ids = [c.get("id") for c in customers if c.get("id")]
                all_transactions = await db.payment_transactions.find(
                    {"customer_id": {"$in": page_cust_ids}}, {"_id": 0, "customer_id": 1, "amount": 1}
                ).to_list(10000)
                txn_by_customer = {}
                for txn in all_transactions:
                    cid = txn.get("customer_id")
                    if cid not in txn_by_customer:
                        txn_by_customer[cid] = []
                    txn_by_customer[cid].append(txn)
                
                overdue_customers = []
                for cust in customers:
                    cust_id = cust.get("id")
                    cust_total_price = cust.get("total_price", 0) or 0
                    cust_booking_amount = cust.get("booking_amount", 0) or 0
                    
                    expected_amount = (cust_total_price * cumulative_percentage) / 100
                    
                    cust_txns = txn_by_customer.get(cust_id, [])
                    txn_total = sum(t.get("amount", 0) or 0 for t in cust_txns)
                    # Total received = transactions only (booking_amount is already in transactions)
                    total_received = txn_total
                    
                    overdue_amt = expected_amount - total_received
                    if overdue_amt > 0:
                        cust['_overdue_amount'] = round(overdue_amt, 2)
                        overdue_customers.append(cust)
                
                # Sort by overdue amount (highest first)
                overdue_customers.sort(key=lambda x: x.get('_overdue_amount', 0), reverse=True)
                customers = overdue_customers[:limit]
        else:
            customers = []  # No stage set, no overdue customers
    
    total = await db.customers.count_documents(query) if agreement_filter not in ["upcoming_due", "overdue"] else len(customers)
    
    return {"customers": customers, "total": total}

@api_router.get("/customers/{customer_id}")
async def get_customer(customer_id: str, user: dict = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return customer

@api_router.put("/customers/{customer_id}")
async def update_customer(customer_id: str, updates: Dict[str, Any], user: dict = Depends(get_current_user)):
    # Accounts role can only update agreement_status, not other customer details
    if user['role'] == 'accounts':
        allowed_fields = {'agreement_status'}
        if not set(updates.keys()).issubset(allowed_fields):
            raise HTTPException(status_code=403, detail="Accounts role can only update agreement status")
    
    # PROTECT: Never allow overwriting one-time booking details via general update
    protected_booking_fields = {
        'booking_amount', 'booking_date', 'transaction_date',
        'transaction_bank', 'transaction_details'
    }
    for field in protected_booking_fields:
        updates.pop(field, None)
    
    updates['updated_at'] = datetime.now(timezone.utc).isoformat()
    result = await db.customers.update_one({"id": customer_id}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    await log_activity(user['id'], user['name'], "update", "customer", customer_id, "Updated customer")
    return {"message": "Customer updated"}

@api_router.put("/customers/{customer_id}/booking-details")
async def update_booking_details(customer_id: str, updates: Dict[str, Any], user: dict = Depends(check_role([UserRole.ADMIN]))):
    """Admin-only endpoint to update booking details (finance_type, finance_bank, booking_amount, booking_date)."""
    allowed = {'finance_type', 'finance_bank', 'booking_amount', 'booking_date'}
    filtered = {k: v for k, v in updates.items() if k in allowed}
    if not filtered:
        raise HTTPException(status_code=400, detail="No valid booking fields provided")
    
    if 'booking_amount' in filtered:
        filtered['booking_amount'] = float(filtered['booking_amount']) if filtered['booking_amount'] else 0
    
    filtered['updated_at'] = datetime.now(timezone.utc).isoformat()
    result = await db.customers.update_one({"id": customer_id}, {"$set": filtered})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    await log_activity(user['id'], user['name'], "update", "customer", customer_id, f"Updated booking details: {list(filtered.keys())}")
    return {"message": "Booking details updated"}

@api_router.delete("/customers/{customer_id}")
async def delete_customer(customer_id: str, user: dict = Depends(check_role([UserRole.ADMIN, UserRole.MANAGER]))):
    result = await db.customers.delete_one({"id": customer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Clean up related data
    await db.payment_schedules.delete_many({"customer_id": customer_id})
    await db.document_checklists.delete_one({"customer_id": customer_id})
    await db.generated_documents.delete_many({"customer_id": customer_id})
    await db.communication_logs.delete_many({"customer_id": customer_id})
    
    await log_activity(user['id'], user['name'], "delete", "customer", customer_id, "Deleted customer")
    return {"message": "Customer deleted"}

# ==================== PAYMENT SCHEDULE ROUTES ====================
@api_router.post("/payments/schedule")
async def create_payment_schedule(data: PaymentScheduleCreate, user: dict = Depends(get_current_user)):
    existing = await db.payment_schedules.find_one({"customer_id": data.customer_id}, {"_id": 0})
    
    schedule_doc = {
        "id": existing['id'] if existing else str(uuid.uuid4()),
        "customer_id": data.customer_id,
        "items": [item.model_dump() for item in data.items],
        "created_at": existing['created_at'] if existing else datetime.now(timezone.utc).isoformat()
    }
    
    if existing:
        await db.payment_schedules.update_one({"customer_id": data.customer_id}, {"$set": schedule_doc})
    else:
        await db.payment_schedules.insert_one(schedule_doc)
    
    await log_activity(user['id'], user['name'], "update", "payment_schedule", data.customer_id, "Updated payment schedule")
    return {"message": "Payment schedule saved", "schedule": schedule_doc}

@api_router.get("/payments/schedule/{customer_id}")
async def get_payment_schedule(customer_id: str, user: dict = Depends(get_current_user)):
    schedule = await db.payment_schedules.find_one({"customer_id": customer_id}, {"_id": 0})
    if not schedule:
        return {"customer_id": customer_id, "items": []}
    return schedule

@api_router.put("/payments/item/{customer_id}/{item_id}")
async def update_payment_item(customer_id: str, item_id: str, updates: Dict[str, Any], user: dict = Depends(get_current_user)):
    schedule = await db.payment_schedules.find_one({"customer_id": customer_id}, {"_id": 0})
    if not schedule:
        raise HTTPException(status_code=404, detail="Payment schedule not found")
    
    for item in schedule['items']:
        if item['id'] == item_id:
            item.update(updates)
            break
    
    await db.payment_schedules.update_one({"customer_id": customer_id}, {"$set": {"items": schedule['items']}})
    
    # Auto-calculate total_received based on paid items
    total_received = 0
    for item in schedule['items']:
        if item.get('payment_status') == 'paid':
            total_received += item.get('amount', 0)
        elif item.get('payment_status') == 'partial':
            # For partial payments, count 50% of the amount (or you can add a partial_amount field later)
            total_received += item.get('amount', 0) * 0.5
    
    # Get customer's total price to calculate percentages
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0, "total_price": 1})
    total_price = customer.get('total_price', 0) if customer else 0
    
    # Calculate payment percentages
    payment_received_percentage = round((total_received / total_price * 100), 2) if total_price > 0 else 0
    payment_pending_percentage = round(100 - payment_received_percentage, 2)
    balance_amount = round(total_price - total_received, 2)
    
    # Update customer's payment tracking fields
    await db.customers.update_one(
        {"id": customer_id},
        {"$set": {
            "total_received": round(total_received, 2),
            "balance_amount": balance_amount,
            "payment_received_percentage": payment_received_percentage,
            "payment_pending_percentage": payment_pending_percentage,
            "updated_at": datetime.now(timezone.utc)
        }}
    )
    
    await log_activity(user['id'], user['name'], "update", "payment_item", item_id, f"Updated payment status - Total Received: {total_received}")
    
    # Return updated payment tracking info
    return {
        "message": "Payment item updated",
        "total_received": round(total_received, 2),
        "balance_amount": balance_amount,
        "payment_received_percentage": payment_received_percentage,
        "payment_pending_percentage": payment_pending_percentage
    }

@api_router.get("/payments/overview")
async def get_payments_overview(user: dict = Depends(get_current_user)):
    today = datetime.now(timezone.utc).date()
    week_end = today + timedelta(days=7)
    
    schedules = await db.payment_schedules.find({}, {"_id": 0, "customer_id": 1, "items": 1}).to_list(1000)
    
    # Fix N+1 query: Fetch all relevant customers in one query
    customer_ids = list(set(s.get('customer_id') for s in schedules if s.get('customer_id')))
    customers_list = await db.customers.find(
        {"id": {"$in": customer_ids}}, 
        {"_id": 0, "id": 1, "name": 1, "customer_id": 1, "unit_number": 1}
    ).to_list(1000)
    customers_dict = {c['id']: c for c in customers_list}
    
    pending = []
    overdue = []
    upcoming = []
    
    for schedule in schedules:
        customer = customers_dict.get(schedule.get('customer_id'))
        for item in schedule.get('items', []):
            if item['payment_status'] == 'paid':
                continue
            
            try:
                due_date = datetime.strptime(item['due_date'], "%Y-%m-%d").date()
            except (ValueError, TypeError):
                continue
            
            item_data = {**item, "customer_name": customer.get('name', 'N/A') if customer else 'N/A', 
                        "customer_ref": customer.get('customer_id', '') if customer else '',
                        "unit_number": customer.get('unit_number', '') if customer else ''}
            
            if due_date < today:
                overdue.append(item_data)
            elif due_date <= week_end:
                upcoming.append(item_data)
            else:
                pending.append(item_data)
    
    return {"pending": pending, "overdue": overdue, "upcoming": upcoming}

# ==================== PAYMENT TRANSACTIONS ====================

async def auto_generate_booking_transaction(customer_id: str, customer: dict, created_by: str = "system"):
    """
    Auto-generate a booking transaction if the customer has a booking_amount
    that isn't already covered by existing booking-stage transactions.
    This ensures total_received (from transactions) always includes the booking payment.
    """
    booking_amount = customer.get("booking_amount", 0) or 0
    if booking_amount <= 0:
        return

    # Check existing booking-stage transactions (check both legacy 'transaction_type' and new 'transaction_stage' fields)
    existing_txns = await db.payment_transactions.find(
        {"customer_id": customer_id, "$or": [
            {"transaction_stage": "booking"},
            {"transaction_type": "booking"}
        ]},
        {"_id": 0, "amount": 1}
    ).to_list(1000)
    existing_booking_sum = sum(t.get("amount", 0) or 0 for t in existing_txns)

    if existing_booking_sum >= booking_amount:
        return  # Already covered

    # Create transaction for the shortfall
    shortfall = booking_amount - existing_booking_sum
    booking_date = customer.get("booking_date") or datetime.now(timezone.utc).strftime("%Y-%m-%d")
    txn_bank = customer.get("transaction_bank", "") or ""
    txn_ref = customer.get("transaction_details", "") or ""
    unit = customer.get("unit_number", "") or ""
    name = customer.get("name", "") or ""

    new_txn = PaymentTransaction(
        customer_id=customer_id,
        transaction_stage=TransactionStage.BOOKING,
        transaction_date=booking_date,
        bank_name=txn_bank,
        transaction_number=txn_ref,
        amount=shortfall,
        notes=f"Auto-generated from booking amount. Flat: {unit}, Client: {name}".strip(),
    )
    await db.payment_transactions.insert_one(new_txn.model_dump())

    # Recalculate customer's total_received from all transactions
    all_txns = await db.payment_transactions.find(
        {"customer_id": customer_id}, {"_id": 0, "amount": 1}
    ).to_list(1000)
    total_received = sum(t.get("amount", 0) or 0 for t in all_txns)
    total_price = customer.get("total_price", 0) or 0
    balance_amount = total_price - total_received

    await db.customers.update_one(
        {"id": customer_id},
        {"$set": {
            "total_received": total_received,
            "balance_amount": balance_amount,
            "payment_received_percentage": round((total_received / total_price) * 100, 2) if total_price > 0 else 0,
            "payment_pending_percentage": round((balance_amount / total_price) * 100, 2) if total_price > 0 else 100
        }}
    )
    logger.info(f"Auto-generated booking transaction of {shortfall} for customer {customer_id} ({name})")

@api_router.get("/transactions/{customer_id}")
async def get_transactions(customer_id: str, user: dict = Depends(get_current_user)):
    """Get all transactions for a customer"""
    transactions = await db.payment_transactions.find(
        {"customer_id": customer_id}, {"_id": 0}
    ).sort("transaction_date", -1).to_list(1000)
    # Normalize: ensure transaction_stage is present (fallback from transaction_type)
    for t in transactions:
        if not t.get("transaction_stage") and t.get("transaction_type"):
            t["transaction_stage"] = t["transaction_type"]
    return transactions

@api_router.post("/transactions/{customer_id}")
async def create_transaction(customer_id: str, transaction: PaymentTransactionCreate, user: dict = Depends(get_current_user)):
    """Create a new transaction"""
    # Verify customer exists - check both id and customer_id fields
    customer = await db.customers.find_one({"$or": [{"id": customer_id}, {"customer_id": customer_id}]})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    new_transaction = PaymentTransaction(
        customer_id=customer_id,
        transaction_stage=transaction.transaction_stage,
        transaction_date=transaction.transaction_date,
        bank_name=transaction.bank_name,
        transaction_number=transaction.transaction_number,
        amount=transaction.amount,
        notes=transaction.notes
    )
    
    await db.payment_transactions.insert_one(new_transaction.model_dump())
    
    # Update customer's total_received and balance_amount
    all_transactions = await db.payment_transactions.find({"customer_id": customer_id}, {"_id": 0, "amount": 1}).to_list(1000)
    total_received = sum(t.get('amount', 0) or 0 for t in all_transactions)
    total_price = customer.get('total_price', 0) or 0
    balance_amount = total_price - total_received
    
    await db.customers.update_one(
        {"$or": [{"id": customer_id}, {"customer_id": customer_id}]},
        {"$set": {
            "total_received": total_received,
            "balance_amount": balance_amount,
            "payment_received_percentage": round((total_received / total_price) * 100, 2) if total_price > 0 else 0,
            "payment_pending_percentage": round((balance_amount / total_price) * 100, 2) if total_price > 0 else 100
        }}
    )
    
    await log_activity(user['id'], user['name'], "create", "transaction", new_transaction.id, f"Created transaction for customer {customer_id}")
    
    return {"message": "Transaction created", "transaction": new_transaction.model_dump()}

@api_router.put("/transactions/{customer_id}/{transaction_id}")
async def update_transaction(customer_id: str, transaction_id: str, transaction: PaymentTransactionCreate, user: dict = Depends(get_current_user)):
    """Update an existing transaction"""
    existing = await db.payment_transactions.find_one({"id": transaction_id, "customer_id": customer_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    update_data = {
        "transaction_stage": transaction.transaction_stage.value if hasattr(transaction.transaction_stage, 'value') else transaction.transaction_stage,
        "transaction_date": transaction.transaction_date,
        "bank_name": transaction.bank_name,
        "transaction_number": transaction.transaction_number,
        "amount": transaction.amount,
        "notes": transaction.notes,
        "updated_at": datetime.now(timezone.utc)
    }
    
    await db.payment_transactions.update_one(
        {"id": transaction_id},
        {"$set": update_data}
    )
    
    # Update customer's total_received and balance_amount
    customer = await db.customers.find_one({"$or": [{"id": customer_id}, {"customer_id": customer_id}]})
    if customer:
        all_transactions = await db.payment_transactions.find({"customer_id": customer_id}, {"_id": 0, "amount": 1}).to_list(1000)
        total_received = sum(t.get('amount', 0) or 0 for t in all_transactions)
        total_price = customer.get('total_price', 0) or 0
        balance_amount = total_price - total_received
        
        await db.customers.update_one(
            {"$or": [{"id": customer_id}, {"customer_id": customer_id}]},
            {"$set": {
                "total_received": total_received,
                "balance_amount": balance_amount,
                "payment_received_percentage": round((total_received / total_price) * 100, 2) if total_price > 0 else 0,
                "payment_pending_percentage": round((balance_amount / total_price) * 100, 2) if total_price > 0 else 100
            }}
        )
    
    await log_activity(user['id'], user['name'], "update", "transaction", transaction_id, f"Updated transaction for customer {customer_id}")
    
    return {"message": "Transaction updated"}

@api_router.delete("/transactions/{customer_id}/{transaction_id}")
async def delete_transaction(customer_id: str, transaction_id: str, user: dict = Depends(get_current_user)):
    """Delete a transaction - restricted for accounts role"""
    # Accounts role cannot delete transactions
    if user['role'] == 'accounts':
        raise HTTPException(status_code=403, detail="Accounts role cannot delete transactions")
    
    result = await db.payment_transactions.delete_one({"id": transaction_id, "customer_id": customer_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transaction not found")
    
    # Update customer's total_received and balance_amount
    customer = await db.customers.find_one({"$or": [{"id": customer_id}, {"customer_id": customer_id}]})
    if customer:
        all_transactions = await db.payment_transactions.find({"customer_id": customer_id}, {"_id": 0, "amount": 1}).to_list(1000)
        total_received = sum(t.get('amount', 0) or 0 for t in all_transactions)
        total_price = customer.get('total_price', 0) or 0
        balance_amount = total_price - total_received
        
        await db.customers.update_one(
            {"$or": [{"id": customer_id}, {"customer_id": customer_id}]},
            {"$set": {
                "total_received": total_received,
                "balance_amount": balance_amount,
                "payment_received_percentage": round((total_received / total_price) * 100, 2) if total_price > 0 else 0,
                "payment_pending_percentage": round((balance_amount / total_price) * 100, 2) if total_price > 0 else 100
            }}
        )
    
    await log_activity(user['id'], user['name'], "delete", "transaction", transaction_id, f"Deleted transaction for customer {customer_id}")
    
    return {"message": "Transaction deleted"}

@api_router.get("/transactions/{customer_id}/export-html")
async def export_transactions_html(customer_id: str, user: dict = Depends(get_current_user)):
    """Generate HTML for transaction details PDF export"""
    customer = await db.customers.find_one(
        {"$or": [{"id": customer_id}, {"customer_id": customer_id}]}, {"_id": 0}
    )
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Query transactions using both possible customer_id values (UUID and RRL-XXXXX)
    cust_uuid = customer.get('id', '')
    cust_display_id = customer.get('customer_id', '')
    possible_ids = list(set(filter(None, [customer_id, cust_uuid, cust_display_id])))
    
    transactions = await db.payment_transactions.find(
        {"customer_id": {"$in": possible_ids}}, {"_id": 0}
    ).sort("transaction_date", 1).to_list(1000)
    
    for t in transactions:
        if not t.get("transaction_stage") and t.get("transaction_type"):
            t["transaction_stage"] = t["transaction_type"]
    
    from documents.templates.common import get_logo_img_tag, COMPANY_NAME, format_customer_names
    
    def fmt_inr(amount):
        amount = float(amount) if amount else 0
        int_part = int(amount)
        s = str(int_part)
        if len(s) > 3:
            result = s[-3:]
            s = s[:-3]
            while s:
                result = s[-2:] + ',' + result
                s = s[:-2]
        else:
            result = s
        return f"\u20b9{result}"
    
    total_received = sum(float(t.get('amount', 0) or 0) for t in transactions)
    total_price = float(customer.get('total_price', 0) or 0)
    balance = total_price - total_received
    
    customer_names = format_customer_names(customer)
    co_applicant_row = ""
    if customer.get('co_applicant_name'):
        co_applicant_row = f'''
            <div class="info-item">
                <div class="info-label">Co-Applicant</div>
                <div class="info-value">{customer.get('co_applicant_name', '')}</div>
            </div>
            <div class="info-item">
                <div class="info-label">Co-Applicant Phone</div>
                <div class="info-value">{customer.get('co_applicant_phone', '') or '-'}</div>
            </div>'''
    
    txn_rows = ""
    for i, txn in enumerate(transactions, 1):
        amount = txn.get('amount', 0) or 0
        stage = (txn.get('transaction_stage', '') or 'Payment').replace('_', ' ').title()
        txn_date = txn.get('transaction_date', '-')
        bank = txn.get('bank_name', '-') or '-'
        txn_no = txn.get('transaction_number', '-') or '-'
        notes = txn.get('notes', '') or ''
        txn_rows += f'''
        <tr>
            <td style="text-align: center;">{i}</td>
            <td>{txn_date}</td>
            <td>{stage}</td>
            <td>{bank}</td>
            <td>{txn_no}</td>
            <td style="text-align: right; font-weight: 500;">{fmt_inr(amount)}</td>
            <td>{notes}</td>
        </tr>'''
    
    if not txn_rows:
        txn_rows = '<tr><td colspan="7" style="text-align: center; padding: 20px; color: #666;">No transactions recorded</td></tr>'
    
    html = f'''<!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@300;400;500;700&display=swap');
            * {{ margin: 0; padding: 0; box-sizing: border-box; }}
            body {{ font-family: 'Roboto', sans-serif; padding: 30px; color: #1A1A1A; background: #fff; }}
            .container {{ max-width: 900px; margin: 0 auto; }}
            .header {{ display: flex; justify-content: space-between; align-items: center; border-bottom: 3px solid #D4AF37; padding-bottom: 20px; margin-bottom: 25px; }}
            .logo-section {{ display: flex; align-items: center; gap: 15px; }}
            .logo img {{ width: 100px; height: auto; }}
            .company-name {{ font-size: 20px; font-weight: 700; color: #1A1A1A; }}
            .company-tagline {{ font-size: 11px; color: #666; }}
            .document-title {{ background: #1A1A1A; color: #D4AF37; padding: 10px 20px; border-radius: 4px; font-weight: 500; font-size: 13px; text-transform: uppercase; }}
            .customer-info {{ background: #fafafa; padding: 15px; border-radius: 8px; margin-bottom: 20px; border-left: 4px solid #D4AF37; display: grid; grid-template-columns: 1fr 1fr; gap: 10px; }}
            .info-item {{ padding: 5px 0; }}
            .info-label {{ color: #666; font-size: 11px; }}
            .info-value {{ font-weight: 500; font-size: 12px; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
            th {{ background: #1A1A1A; color: #D4AF37; padding: 10px 8px; text-align: left; font-size: 11px; font-weight: 500; }}
            td {{ padding: 10px 8px; border-bottom: 1px solid #e0e0e0; font-size: 11px; }}
            tr:nth-child(even) {{ background: #fafafa; }}
            .summary {{ margin-top: 20px; display: grid; grid-template-columns: repeat(3, 1fr); gap: 15px; }}
            .summary-box {{ padding: 15px; border-radius: 8px; text-align: center; }}
            .summary-box.total {{ background: #1A1A1A; color: #D4AF37; }}
            .summary-box.received {{ background: #e8f5e9; border: 1px solid #28a745; }}
            .summary-box.balance {{ background: #fff3e0; border: 1px solid #D4AF37; }}
            .summary-label {{ font-size: 10px; text-transform: uppercase; }}
            .summary-value {{ font-size: 18px; font-weight: 700; margin-top: 5px; }}
            .footer {{ margin-top: 30px; padding-top: 15px; border-top: 2px solid #D4AF37; text-align: center; font-size: 10px; color: #666; }}
        </style>
    </head>
    <body>
        <div class="container">
            <div class="header">
                <div class="logo-section">
                    <div class="logo">{get_logo_img_tag(100)}</div>
                    <div>
                        <div class="company-name">{COMPANY_NAME}</div>
                        <div class="company-tagline">Beyond homes. A lifestyle</div>
                    </div>
                </div>
                <div class="document-title">Transaction Details</div>
            </div>
            
            <div class="customer-info">
                <div class="info-item">
                    <div class="info-label">Customer Name</div>
                    <div class="info-value">{customer.get('name', '-')}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">Phone</div>
                    <div class="info-value">{customer.get('phone', '-')}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">Project</div>
                    <div class="info-value">{customer.get('project', '-')}</div>
                </div>
                <div class="info-item">
                    <div class="info-label">Unit Number</div>
                    <div class="info-value">{customer.get('tower', '')}-{customer.get('unit_number', '-')}</div>
                </div>
                {co_applicant_row}
            </div>
            
            <table>
                <thead>
                    <tr>
                        <th style="width: 4%;">#</th>
                        <th style="width: 12%;">Date</th>
                        <th style="width: 16%;">Stage</th>
                        <th style="width: 16%;">Bank</th>
                        <th style="width: 18%;">Transaction No.</th>
                        <th style="width: 14%; text-align: right;">Amount</th>
                        <th style="width: 20%;">Notes</th>
                    </tr>
                </thead>
                <tbody>
                    {txn_rows}
                </tbody>
            </table>
            
            <div class="summary">
                <div class="summary-box total">
                    <div class="summary-label">Total Property Value</div>
                    <div class="summary-value">{fmt_inr(total_price)}</div>
                </div>
                <div class="summary-box received">
                    <div class="summary-label">Total Received</div>
                    <div class="summary-value" style="color: #28a745;">{fmt_inr(total_received)}</div>
                </div>
                <div class="summary-box balance">
                    <div class="summary-label">Balance Pending</div>
                    <div class="summary-value" style="color: #D4AF37;">{fmt_inr(balance)}</div>
                </div>
            </div>
            
            <div class="footer">
                <p>{COMPANY_NAME} | www.rrlbuildersanddevelopers.com</p>
                <p>Generated on {datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M")} UTC</p>
            </div>
        </div>
    </body>
    </html>'''
    
    return {"content": html, "customer_name": customer.get('name', 'Customer')}

# ==================== PRICE CALCULATOR ====================
# Default payment schedule template (from Excel)
DEFAULT_PAYMENT_SCHEDULE = [
    {"installment_name": "Initial Booking Amount", "percentage": 10, "milestone": "booking", "description": "Balance booking amount (To be paid within 10 days of Booking)"},
    {"installment_name": "Post Excavation of Agreement", "percentage": 10, "milestone": "agreement", "description": "To be paid within 10 days of Booking"},
    {"installment_name": "On Completion of Foundation", "percentage": 10, "milestone": "foundation", "description": ""},
    {"installment_name": "On Completion of Podium Slab", "percentage": 10, "milestone": "podium", "description": ""},
    {"installment_name": "Upon Completion of 2nd Floor Roof Slab", "percentage": 5, "milestone": "2nd_floor", "description": ""},
    {"installment_name": "Upon Completion of 6th Floor Roof Slab", "percentage": 5, "milestone": "6th_floor", "description": ""},
    {"installment_name": "Upon Completion of 10th Floor Roof Slab", "percentage": 5, "milestone": "10th_floor", "description": ""},
    {"installment_name": "Upon Completion of 14th Floor Roof Slab", "percentage": 5, "milestone": "14th_floor", "description": ""},
    {"installment_name": "Upon Completion of 18th Floor Roof Slab", "percentage": 5, "milestone": "18th_floor", "description": ""},
    {"installment_name": "Upon Completion of 22nd Floor Roof Slab", "percentage": 5, "milestone": "22nd_floor", "description": ""},
    {"installment_name": "Upon Completion of Top Roof Slab", "percentage": 10, "milestone": "top_roof", "description": ""},
    {"installment_name": "Upon Completion of Flooring of Particular Property", "percentage": 10, "milestone": "flooring", "description": ""},
    {"installment_name": "Upon Handover or Possession of Particular Property or Registration of Absolute Sale for Particular Property, whichever is Earlier", "percentage": 10, "milestone": "handover", "description": ""},
]

@api_router.post("/calculator/price", response_model=PriceResult)
async def calculate_price(data: PriceCalculation):
    """
    Calculate total flat value with all charges
    Formula: (Rate/sqft × Saleable Area) + Club House + Additional Charges + Labour Cess + GST
    """
    # Base price = Rate × Saleable Area
    base_price = data.rate_per_sqft * data.saleable_area
    
    # Club house charges (editable, default Rs. 2,00,000)
    club_house = data.club_house_charges if data.include_club_house else 0
    
    # Additional manual charges
    additional_charges = data.additional_charges or 0
    
    # Subtotal before taxes
    subtotal = base_price + club_house + additional_charges
    
    # Labour cess (0.70% of subtotal)
    labour_cess = subtotal * (data.labour_cess_percentage / 100)
    
    # GST (5% of subtotal)
    gst_amount = subtotal * (data.gst_percentage / 100)
    
    # Total flat value
    total_flat_value = subtotal + labour_cess + gst_amount
    
    # UDS calculation
    uds = data.saleable_area * 0.495046
    
    return PriceResult(
        unit_number=data.unit_number,
        unit_type=data.unit_type,
        floor_number=data.floor_number,
        saleable_area=data.saleable_area,
        rate_per_sqft=data.rate_per_sqft,
        base_price=round(base_price, 2),
        club_house_charges=round(club_house, 2),
        additional_charges=round(additional_charges, 2),
        subtotal_before_taxes=round(subtotal, 2),
        labour_cess=round(labour_cess, 2),
        gst_amount=round(gst_amount, 2),
        total_flat_value=round(total_flat_value, 2),
        uds=round(uds, 2)
    )

@api_router.post("/calculator/disbursement", response_model=DisbursementResult)
async def calculate_disbursement(data: DisbursementCalculation):
    """
    Calculate disbursement amount
    Formula: Total Flat Value × Disbursement Percentage
    """
    disbursement_amount = data.total_flat_value * (data.disbursement_percentage / 100)
    
    return DisbursementResult(
        total_flat_value=round(data.total_flat_value, 2),
        disbursement_percentage=data.disbursement_percentage,
        disbursement_amount=round(disbursement_amount, 2)
    )

@api_router.post("/calculator/payment-tracking", response_model=PaymentTrackingResult)
async def calculate_payment_tracking(total_flat_value: float, total_received: float):
    """
    Calculate payment tracking metrics
    - Balance = Total - Received
    - Received % = (Received / Total) × 100
    - Pending % = 100 - Received %
    """
    balance = total_flat_value - total_received
    received_percentage = (total_received / total_flat_value * 100) if total_flat_value > 0 else 0
    pending_percentage = 100 - received_percentage
    
    return PaymentTrackingResult(
        total_flat_value=round(total_flat_value, 2),
        total_received=round(total_received, 2),
        balance_amount=round(balance, 2),
        payment_received_percentage=round(received_percentage, 2),
        payment_pending_percentage=round(pending_percentage, 2)
    )

@api_router.get("/calculator/payment-schedule-template")
async def get_payment_schedule_template(total_amount: float = 0):
    """Get the default payment schedule template with calculated amounts"""
    schedule = []
    cumulative = 0
    for item in DEFAULT_PAYMENT_SCHEDULE:
        amount = total_amount * (item["percentage"] / 100) if total_amount > 0 else 0
        cumulative += amount
        schedule.append({
            **item,
            "amount": round(amount, 2),
            "cumulative": round(cumulative, 2)
        })
    return schedule

@api_router.post("/calculator/generate-schedule/{customer_id}")
async def generate_payment_schedule_for_customer(customer_id: str, user: dict = Depends(get_current_user)):
    """Auto-generate payment schedule based on customer's total price"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    total_amount = customer.get("total_price", 0)
    if total_amount <= 0:
        raise HTTPException(status_code=400, detail="Customer has no total price set")
    
    items = []
    cumulative = 0
    for item in DEFAULT_PAYMENT_SCHEDULE:
        amount = total_amount * (item["percentage"] / 100)
        cumulative += amount
        items.append({
            "id": str(uuid.uuid4()),
            "installment_name": item["installment_name"],
            "milestone": item["milestone"],
            "description": item.get("description", ""),
            "percentage": item["percentage"],
            "amount": round(amount, 2),
            "cumulative": round(cumulative, 2),
            "due_date": "",
            "payment_status": "pending",
            "payment_date": None
        })
    
    schedule_doc = {
        "id": str(uuid.uuid4()),
        "customer_id": customer_id,
        "items": items,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    # Upsert the schedule
    await db.payment_schedules.update_one(
        {"customer_id": customer_id},
        {"$set": schedule_doc},
        upsert=True
    )
    
    await log_activity(user['id'], user['name'], "generate", "payment_schedule", customer_id, "Auto-generated payment schedule")
    
    return {"message": "Payment schedule generated", "schedule": schedule_doc}

# ==================== DOCUMENT TEMPLATES ====================
@api_router.get("/templates")
async def get_templates(user: dict = Depends(get_current_user)):
    templates = await db.document_templates.find({}, {"_id": 0}).to_list(100)
    return templates

@api_router.post("/templates")
async def create_template(template: DocumentTemplate, user: dict = Depends(check_role([UserRole.ADMIN, UserRole.MANAGER]))):
    doc = template.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await db.document_templates.insert_one(doc)
    await log_activity(user['id'], user['name'], "create", "template", template.id, f"Created template {template.name}")
    return {"message": "Template created", "id": template.id}

@api_router.put("/templates/{template_id}")
async def update_template(template_id: str, updates: Dict[str, Any], user: dict = Depends(check_role([UserRole.ADMIN, UserRole.MANAGER]))):
    updates['updated_at'] = datetime.now(timezone.utc).isoformat()
    result = await db.document_templates.update_one({"id": template_id}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Template not found")
    
    await log_activity(user['id'], user['name'], "update", "template", template_id, "Updated template")
    return {"message": "Template updated"}

# ==================== DOCUMENT GENERATION ====================
@api_router.post("/documents/generate")
async def generate_document(data: DocumentGenerate, user: dict = Depends(get_current_user)):
    customer = await db.customers.find_one({"id": data.customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # For Sales Agreement, use the dedicated generator function
    if data.doc_type == DocumentType.SALES_AGREEMENT:
        # Get payment schedule
        schedule = await db.payment_schedules.find_one({"customer_id": data.customer_id}, {"_id": 0})
        schedule_items = schedule.get('items', []) if schedule else []
        
        # Get transaction records
        transactions = await db.payment_transactions.find(
            {"customer_id": data.customer_id}, {"_id": 0}
        ).sort("transaction_date", 1).to_list(1000)
        
        content = generate_sales_agreement_html(customer, schedule_items, transactions)
    elif data.doc_type == DocumentType.PRICE_BREAKUP:
        content = generate_price_breakup_html(customer)
    elif data.doc_type == DocumentType.COST_BREAKUP:
        content = generate_cost_breakup_html(customer)
    elif data.doc_type == DocumentType.ALLOTMENT_LETTER:
        content = generate_allotment_letter_html(customer)
    elif data.doc_type == DocumentType.PAYMENT_SCHEDULE:
        # Get transaction records for payment schedule
        transactions = await db.payment_transactions.find(
            {"customer_id": data.customer_id}, {"_id": 0}
        ).sort("transaction_date", 1).to_list(1000)
        content = generate_payment_schedule_pdf_html(customer, transactions)
    elif data.doc_type == DocumentType.NOC_HDFC:
        content = generate_noc_hdfc_html(customer)
    elif data.doc_type == DocumentType.NOC_BOB:
        content = generate_noc_bob_html(customer)
    elif data.doc_type == DocumentType.NOC_TATA:
        content = generate_noc_tata_html(customer)
    elif data.doc_type == DocumentType.DEMAND_LETTER:
        # Fetch transactions
        transactions = await db.payment_transactions.find(
            {"customer_id": data.customer_id}, {"_id": 0}
        ).sort("transaction_date", 1).to_list(1000)
        # Fetch current construction stage from settings
        settings = await db.settings.find_one({"type": "payment_stage"}, {"_id": 0})
        stage_info = {}
        if settings and settings.get("current_stage"):
            stage_key = settings.get("current_stage")
            stage_info = next((s for s in PAYMENT_STAGES if s["key"] == stage_key), {})
        content = generate_demand_letter_html(customer, transactions, stage_info)
    else:
        # For other document types, use template-based generation
        template = await db.document_templates.find_one({"doc_type": data.doc_type.value}, {"_id": 0})
        if not template:
            template = {"content": get_default_template(data.doc_type)}
        
        content = template['content']
        
        # Format total price with Indian currency format
        total_price = customer.get('total_price', 0)
        total_price_formatted = format_indian_currency(total_price, decimals=False) if total_price else "0"
        
        # Calculate UDS if not present
        uds = customer.get('uds', 0)
        if not uds and customer.get('saleable_area'):
            uds = round(customer.get('saleable_area', 0) * 0.495046, 2)
        
        placeholders = {
            "{customer_name}": customer.get('name', ''),
            "{customer_id}": customer.get('customer_id', ''),
            "{unit_number}": customer.get('unit_number', ''),
            "{tower}": customer.get('tower', ''),
            "{project}": customer.get('project', ''),
            "{total_price}": str(total_price),
            "{total_price_formatted}": total_price_formatted,
            "{saleable_area}": str(customer.get('saleable_area', 0)),
            "{uds}": str(uds),
            "{booking_amount}": str(customer.get('booking_amount', 0)),
            "{booking_date}": customer.get('booking_date', ''),
            "{date}": datetime.now().strftime("%d-%m-%Y"),
            "{father_name}": customer.get('father_name', ''),
            "{pan_number}": customer.get('pan_number', ''),
            "{phone}": customer.get('phone', ''),
            "{email}": customer.get('email', ''),
            "{address}": customer.get('address', ''),
            "{bhk_type}": customer.get('bhk_type', ''),
            "{floor}": str(customer.get('floor', '')),
            "{rate_per_sqft}": str(customer.get('rate_per_sqft', 0)),
            "{base_price}": str(customer.get('base_price', 0)),
            "{gst_amount}": str(customer.get('gst_amount', 0)),
            "{labour_cess}": str(customer.get('labour_cess', 0)),
            "{club_house_charges}": str(customer.get('club_house_charges', 0)),
        }
        
        # Add custom fields
        for key, value in data.custom_fields.items():
            placeholders[f"{{{key}}}"] = value
        
        for placeholder, value in placeholders.items():
            content = content.replace(placeholder, str(value))
    
    # Save generated document
    gen_doc = GeneratedDocument(
        customer_id=data.customer_id,
        doc_type=data.doc_type,
        content=content,
        generated_by=user['id']
    )
    
    doc = gen_doc.model_dump()
    doc['generated_at'] = doc['generated_at'].isoformat()
    await db.generated_documents.insert_one(doc)
    
    await log_activity(user['id'], user['name'], "generate", "document", gen_doc.id, f"Generated {data.doc_type.value}")
    
    return {"message": "Document generated", "document": {**doc, "_id": None}}

@api_router.post("/documents/generate-pdf/{customer_id}")
async def generate_price_breakup_pdf(customer_id: str, user: dict = Depends(get_current_user)):
    """Generate Price Breakup PDF for a customer"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Generate HTML
    html_content = generate_price_breakup_html(customer)
    
    # Store the generated document
    gen_doc = GeneratedDocument(
        customer_id=customer_id,
        doc_type=DocumentType.PRICE_BREAKUP,
        content=html_content,
        generated_by=user['id']
    )
    
    doc = gen_doc.model_dump()
    doc['generated_at'] = doc['generated_at'].isoformat()
    await db.generated_documents.insert_one(doc)
    
    await log_activity(user['id'], user['name'], "generate", "price_breakup_pdf", customer_id, "Generated Price Breakup PDF")
    
    # Return HTML that can be converted to PDF on frontend
    return {
        "message": "Price breakup generated",
        "document_id": gen_doc.id,
        "html_content": html_content,
        "filename": f"RRL_PalmAltezze_PriceBreakup_{customer.get('name', 'Customer').replace(' ', '_')}.pdf"
    }

@api_router.get("/communication/preview-welcome-email/{customer_id}")
async def preview_welcome_email(customer_id: str, user: dict = Depends(get_current_user)):
    """
    Preview Welcome Email with 3 PDF attachments before sending:
    1. Booking Form Preview
    2. Terms & Conditions
    3. Price Breakup
    """
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Generate welcome email HTML
    welcome_html = generate_welcome_email_html(customer)
    
    # Generate all 3 PDF attachment HTMLs
    form_preview_html = generate_booking_form_preview_html(customer)
    terms_conditions_html = generate_terms_and_conditions_html(customer)
    price_breakup_html = generate_price_breakup_html(customer)
    
    customer_name_safe = customer.get('name', 'Customer').replace(' ', '_')
    filename_form = f"RRL_BookingFormPreview_{customer_name_safe}.pdf"
    filename_terms = f"RRL_TermsAndConditions_{customer_name_safe}.pdf"
    filename_price = f"RRL_PriceBreakup_{customer_name_safe}.pdf"
    
    recipient_email = customer.get('email')
    subject = f"Welcome to {customer.get('project', 'RRL Builders')} - Booking Confirmation & Documents"
    
    # Default email body (editable)
    default_body = f"""Hello {customer.get('name', '')},

Greetings from RRL Builders and Developers Pvt Ltd.

It is our distinct pleasure to welcome you to {customer.get('project', 'RRL Palm Altezze')} and to congratulate you on the acquisition of your Residence Flat No. {customer.get('unit_number', '')}.

Please find attached the following documents for your reference:
1. Booking Form Preview - Your submitted booking details
2. Terms & Conditions - Important terms governing your allotment
3. Price Breakup - Detailed price calculation

Your dedicated Relationship Director will connect with you personally to ensure that every interaction with us is seamless and tailored to your expectations."""
    
    return {
        "email_type": "welcome",
        "customer_name": customer.get('name'),
        "recipient_email": recipient_email,
        "subject": subject,
        "body": default_body,
        "email_html": welcome_html,
        # 3 attachments for welcome email
        "attachment_html": form_preview_html,  # Primary attachment - Form Preview
        "attachment_filename": filename_form,
        "attachment_html_2": terms_conditions_html,  # Terms & Conditions
        "attachment_filename_2": filename_terms,
        "attachment_html_3": price_breakup_html,  # Price Breakup
        "attachment_filename_3": filename_price,
        "attachments": [filename_form, filename_terms, filename_price],
        "has_sendgrid": bool(SENDGRID_API_KEY)
    }

@api_router.get("/communication/preview-sales-agreement/{customer_id}")
async def preview_sales_agreement_email(customer_id: str, user: dict = Depends(get_current_user)):
    """
    Preview Sales Agreement Email with Sales Agreement and Price Breakup PDFs before sending.
    """
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Get payment schedule for Sales Agreement
    schedule = await db.payment_schedules.find_one({"customer_id": customer_id}, {"_id": 0})
    schedule_items = schedule.get('items', []) if schedule else []
    
    # Get transaction records for payment schedule
    transactions = await db.payment_transactions.find(
        {"customer_id": customer_id}, {"_id": 0}
    ).sort("transaction_date", 1).to_list(1000)
    
    # Generate Sales Agreement HTML with transactions
    sales_agreement_html = generate_sales_agreement_html(customer, schedule_items, transactions)
    
    # Generate price breakup HTML
    price_breakup_html = generate_price_breakup_html(customer)
    
    recipient_email = customer.get('email')
    subject = f"SALE AGREEMENT DRAFT AND PRICE BREAK UP - {customer.get('unit_number', '')}"
    
    # Default email body
    default_body = f"""Hello {customer.get('name', '')},

Greetings from RRL Builders and Developers Pvt Ltd.

We are delighted to take this process ahead, please find attached draft copy of the sale agreement.

We would like to know the date when you are signing up for sale agreement.

Please review the attached documents:
1. Sale Agreement Draft
2. Price Break Up

Looking forward to your confirmation."""
    
    # Generate email HTML (same format as welcome mail)
    email_html = generate_document_email_html(customer, subject, default_body)
    
    return {
        "email_type": "sales_agreement",
        "customer_name": customer.get('name'),
        "recipient_email": recipient_email,
        "subject": subject,
        "body": default_body,
        "email_html": email_html,
        "attachment_html": sales_agreement_html,
        "attachment_html_2": price_breakup_html,
        "attachment_filename": f"RRL_SaleAgreement_{customer.get('name', 'Customer').replace(' ', '_')}.pdf",
        "attachment_filename_2": f"RRL_PriceBreakup_{customer.get('name', 'Customer').replace(' ', '_')}.pdf",
        "has_sendgrid": bool(SENDGRID_API_KEY)
    }

@api_router.get("/communication/preview-allotment-letter/{customer_id}")
async def preview_allotment_letter_email(customer_id: str, user: dict = Depends(get_current_user)):
    """
    Preview Allotment Letter Email with Allotment Letter PDF before sending.
    """
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Generate Allotment Letter HTML
    allotment_letter_html = generate_allotment_letter_html(customer)
    
    recipient_email = customer.get('email')
    subject = f"ALLOTMENT LETTER - {customer.get('project', 'RRL Palm Altezze')} - Flat No. {customer.get('unit_number', '')}"
    
    # Default email body
    default_body = f"""Hello {customer.get('name', '')},

Greetings from RRL Builders and Developers Pvt Ltd.

We are pleased to confirm your allotment for Flat No. {customer.get('unit_number', '')} in {customer.get('project', 'RRL Palm Altezze')}.

Please find attached your Allotment Letter for your records.

Kindly review the terms and conditions mentioned in the letter and let us know if you have any queries."""
    
    # Generate email HTML
    email_html = generate_document_email_html(customer, subject, default_body)
    
    return {
        "email_type": "allotment_letter",
        "customer_name": customer.get('name'),
        "recipient_email": recipient_email,
        "subject": subject,
        "body": default_body,
        "email_html": email_html,
        "attachment_html": allotment_letter_html,
        "attachment_filename": f"RRL_AllotmentLetter_{customer.get('name', 'Customer').replace(' ', '_')}.pdf",
        "has_sendgrid": bool(SENDGRID_API_KEY)
    }

class EmailSendRequest(BaseModel):
    email_type: str  # welcome, sales_agreement, allotment_letter
    subject: str
    body: str
    recipient_email: Optional[str] = None  # Override customer email if provided
    cc: Optional[str] = None  # CC email address

@api_router.post("/communication/send-document-email/{customer_id}")
async def send_document_email(customer_id: str, data: EmailSendRequest, user: dict = Depends(get_current_user)):
    """
    Unified endpoint to send document emails with attachments.
    Supports: welcome, sales_agreement, allotment_letter
    """
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    recipient_email = data.recipient_email or customer.get('email')
    
    # Generate email HTML based on the edited body
    email_html = generate_document_email_html(customer, data.subject, data.body)
    
    # Generate attachments based on email type
    attachments_data = []
    
    if data.email_type == "welcome":
        price_breakup_html = generate_price_breakup_html(customer)
        attachments_data.append({
            "filename": f"RRL_PriceBreakup_{customer.get('name', 'Customer').replace(' ', '_')}.pdf",
            "html": price_breakup_html,
            "doc_type": DocumentType.PRICE_BREAKUP
        })
    elif data.email_type == "sales_agreement":
        schedule = await db.payment_schedules.find_one({"customer_id": customer_id}, {"_id": 0})
        schedule_items = schedule.get('items', []) if schedule else []
        
        # Get transaction records for payment schedule
        transactions = await db.payment_transactions.find(
            {"customer_id": customer_id}, {"_id": 0}
        ).sort("transaction_date", 1).to_list(1000)
        
        sales_agreement_html = generate_sales_agreement_html(customer, schedule_items, transactions)
        price_breakup_html = generate_price_breakup_html(customer)
        
        attachments_data.append({
            "filename": f"RRL_SaleAgreement_{customer.get('name', 'Customer').replace(' ', '_')}.pdf",
            "html": sales_agreement_html,
            "doc_type": DocumentType.SALES_AGREEMENT
        })
        attachments_data.append({
            "filename": f"RRL_PriceBreakup_{customer.get('name', 'Customer').replace(' ', '_')}.pdf",
            "html": price_breakup_html,
            "doc_type": DocumentType.PRICE_BREAKUP
        })
    elif data.email_type == "allotment_letter":
        allotment_letter_html = generate_allotment_letter_html(customer)
        attachments_data.append({
            "filename": f"RRL_AllotmentLetter_{customer.get('name', 'Customer').replace(' ', '_')}.pdf",
            "html": allotment_letter_html,
            "doc_type": DocumentType.ALLOTMENT_LETTER
        })
    
    # Store generated documents
    for att in attachments_data:
        doc = GeneratedDocument(
            customer_id=customer_id,
            doc_type=att['doc_type'],
            content=att['html'],
            generated_by=user['id']
        )
        doc_dict = doc.model_dump()
        doc_dict['generated_at'] = doc_dict['generated_at'].isoformat()
        await db.generated_documents.insert_one(doc_dict)
    
    # Send via SendGrid if configured
    email_status = "pending"
    
    if SENDGRID_API_KEY:
        try:
            message = Mail(
                from_email=(SENDGRID_FROM_EMAIL, SENDGRID_FROM_NAME),
                to_emails=recipient_email,
                subject=data.subject,
                html_content=email_html
            )
            
            # Add CC if provided
            if hasattr(data, 'cc') and data.cc:
                message.add_cc(data.cc)
            
            # Generate PDFs and add as attachments
            for att in attachments_data:
                try:
                    pdf_bytes = HTML(string=att['html']).write_pdf()
                    encoded_pdf = base64.b64encode(pdf_bytes).decode()
                    
                    attachment = Attachment(
                        FileContent(encoded_pdf),
                        FileName(att['filename']),
                        FileType('application/pdf'),
                        Disposition('attachment')
                    )
                    message.add_attachment(attachment)
                    logger.info(f"Added attachment: {att['filename']}")
                except Exception as pdf_error:
                    logger.error(f"Error generating PDF attachment {att['filename']}: {str(pdf_error)}")
            
            sg = SendGridAPIClient(SENDGRID_API_KEY)
            response = sg.send(message)
            
            if response.status_code in [200, 201, 202]:
                email_status = "sent"
                logger.info(f"{data.email_type} email sent to {recipient_email} with {len(attachments_data)} attachments")
            else:
                email_status = "failed"
                logger.error(f"Failed to send {data.email_type} email to {recipient_email}")
                
        except Exception as e:
            email_status = "error"
            logger.error(f"SendGrid error: {str(e)}")
    else:
        email_status = "simulated"
        logger.info(f"SendGrid not configured - {data.email_type} email simulated for {recipient_email}")
    
    # Log communication
    comm_log = {
        "id": str(uuid.uuid4()),
        "customer_id": customer_id,
        "type": "email",
        "subject": data.subject,
        "message": data.body,
        "status": email_status,
        "email_type": data.email_type,
        "attachments": [att['filename'] for att in attachments_data],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user['id']
    }
    await db.communications.insert_one(comm_log)
    
    await log_activity(user['id'], user['name'], "send", "email", customer_id, f"Sent {data.email_type} email to {recipient_email}")
    
    return {
        "message": f"{data.email_type.replace('_', ' ').title()} email sent successfully",
        "status": email_status,
        "recipient": recipient_email,
        "attachments": [att['filename'] for att in attachments_data]
    }

@api_router.post("/communication/send-welcome-email/{customer_id}")
async def send_welcome_email(customer_id: str, user: dict = Depends(get_current_user)):
    """
    Send Welcome Email with 3 PDF attachments via SendGrid:
    1. Booking Form Preview - Shows all submitted form data
    2. Terms & Conditions - Allotment letter terms
    3. Price Breakup - Detailed price calculation
    """
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Generate welcome email HTML
    welcome_html = generate_welcome_email_html(customer)
    
    # Generate all PDF HTMLs
    price_breakup_html = generate_price_breakup_html(customer)
    form_preview_html = generate_booking_form_preview_html(customer)
    terms_conditions_html = generate_terms_and_conditions_html(customer)
    
    # Store the welcome email document
    welcome_doc = GeneratedDocument(
        customer_id=customer_id,
        doc_type=DocumentType.WELCOME_LETTER,
        content=welcome_html,
        generated_by=user['id']
    )
    
    welcome_doc_dict = welcome_doc.model_dump()
    welcome_doc_dict['generated_at'] = welcome_doc_dict['generated_at'].isoformat()
    await db.generated_documents.insert_one(welcome_doc_dict)
    
    # Store price breakup document
    price_doc = GeneratedDocument(
        customer_id=customer_id,
        doc_type=DocumentType.PRICE_BREAKUP,
        content=price_breakup_html,
        generated_by=user['id']
    )
    
    price_doc_dict = price_doc.model_dump()
    price_doc_dict['generated_at'] = price_doc_dict['generated_at'].isoformat()
    await db.generated_documents.insert_one(price_doc_dict)
    
    # File names for attachments
    customer_name_safe = customer.get('name', 'Customer').replace(' ', '_')
    filename_form_preview = f"RRL_BookingFormPreview_{customer_name_safe}.pdf"
    filename_terms = f"RRL_TermsAndConditions_{customer_name_safe}.pdf"
    filename_price = f"RRL_PriceBreakup_{customer_name_safe}.pdf"
    
    recipient_email = customer.get('email')
    subject = f"Welcome to {customer.get('project', 'RRL Builders')} - Booking Confirmation & Documents"
    
    # Send via SendGrid if configured
    email_status = "pending"
    sendgrid_response = None
    attachments_added = []
    
    if SENDGRID_API_KEY:
        try:
            message = Mail(
                from_email=(SENDGRID_FROM_EMAIL, SENDGRID_FROM_NAME),
                to_emails=recipient_email,
                subject=subject,
                html_content=welcome_html
            )
            
            # 1. Generate and attach Booking Form Preview PDF
            try:
                form_pdf_bytes = HTML(string=form_preview_html).write_pdf()
                encoded_form_pdf = base64.b64encode(form_pdf_bytes).decode()
                
                form_attachment = Attachment(
                    FileContent(encoded_form_pdf),
                    FileName(filename_form_preview),
                    FileType('application/pdf'),
                    Disposition('attachment')
                )
                message.add_attachment(form_attachment)
                attachments_added.append(filename_form_preview)
                logger.info(f"Added Booking Form Preview attachment: {filename_form_preview}")
            except Exception as pdf_error:
                logger.error(f"Error generating Form Preview PDF: {str(pdf_error)}")
            
            # 2. Generate and attach Terms & Conditions PDF
            try:
                terms_pdf_bytes = HTML(string=terms_conditions_html).write_pdf()
                encoded_terms_pdf = base64.b64encode(terms_pdf_bytes).decode()
                
                terms_attachment = Attachment(
                    FileContent(encoded_terms_pdf),
                    FileName(filename_terms),
                    FileType('application/pdf'),
                    Disposition('attachment')
                )
                message.add_attachment(terms_attachment)
                attachments_added.append(filename_terms)
                logger.info(f"Added Terms & Conditions attachment: {filename_terms}")
            except Exception as pdf_error:
                logger.error(f"Error generating Terms & Conditions PDF: {str(pdf_error)}")
            
            # 3. Generate and attach Price Breakup PDF
            try:
                price_pdf_bytes = HTML(string=price_breakup_html).write_pdf()
                encoded_price_pdf = base64.b64encode(price_pdf_bytes).decode()
                
                price_attachment = Attachment(
                    FileContent(encoded_price_pdf),
                    FileName(filename_price),
                    FileType('application/pdf'),
                    Disposition('attachment')
                )
                message.add_attachment(price_attachment)
                attachments_added.append(filename_price)
                logger.info(f"Added Price Breakup attachment: {filename_price}")
            except Exception as pdf_error:
                logger.error(f"Error generating Price Breakup PDF: {str(pdf_error)}")
            
            sg = SendGridAPIClient(SENDGRID_API_KEY)
            response = sg.send(message)
            
            if response.status_code in [200, 201, 202]:
                email_status = "sent"
                sendgrid_response = {
                    "status_code": response.status_code, 
                    "body": f"Email sent successfully with {len(attachments_added)} attachments"
                }
                logger.info(f"Welcome email sent to {recipient_email} with {len(attachments_added)} PDFs - Status: {response.status_code}")
            else:
                email_status = "failed"
                sendgrid_response = {"status_code": response.status_code, "error": "Unexpected status code"}
                logger.error(f"Failed to send email to {recipient_email} - Status: {response.status_code}")
                
        except Exception as e:
            email_status = "failed"
            sendgrid_response = {"error": str(e)}
            logger.error(f"SendGrid error sending email to {recipient_email}: {str(e)}")
    else:
        email_status = "mocked (no API key)"
        attachments_added = [filename_form_preview, filename_terms, filename_price]
        logger.warning("SendGrid API key not configured - email not sent")
    
    # Log communication
    log = CommunicationLog(
        customer_id=customer_id,
        channel="email",
        message_type="Welcome Email",
        content=f"""
To: {recipient_email}
Subject: {subject}

[Welcome Email HTML Body]

Attachments:
1. {filename_form_preview} (Booking Form Preview)
2. {filename_terms} (Terms & Conditions)
3. {filename_price} (Price Breakup)
        """,
        status=email_status,
        sent_by=user['id']
    )
    
    log_doc = log.model_dump()
    log_doc['sent_at'] = log_doc['sent_at'].isoformat()
    await db.communication_logs.insert_one(log_doc)
    
    # Update customer stage if still pending
    if customer.get('stage') == 'pending_approval':
        await db.customers.update_one(
            {"id": customer_id},
            {"$set": {"stage": "qualified", "updated_at": datetime.now(timezone.utc).isoformat()}}
        )
    
    await log_activity(user['id'], user['name'], "send", "welcome_email", customer_id, f"Sent welcome email to {recipient_email} with 3 PDFs - Status: {email_status}")
    
    return {
        "message": f"Welcome email {email_status}",
        "welcome_doc_id": welcome_doc.id,
        "price_breakup_doc_id": price_doc.id,
        "email_to": recipient_email,
        "email_status": email_status,
        "attachments": attachments_added,
        "sendgrid_response": sendgrid_response
    }

@api_router.get("/documents/html/{doc_id}")
async def get_document_html(doc_id: str, user: dict = Depends(get_current_user)):
    """Get the HTML content of a generated document for preview/PDF conversion"""
    doc = await db.generated_documents.find_one({"id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    return {
        "id": doc['id'],
        "doc_type": doc['doc_type'],
        "content": doc['content'],
        "generated_at": doc['generated_at']
    }

@api_router.get("/documents/{customer_id}")
async def get_customer_documents(customer_id: str, user: dict = Depends(get_current_user)):
    documents = await db.generated_documents.find({"customer_id": customer_id}, {"_id": 0}).to_list(100)
    return documents

@api_router.put("/documents/{doc_id}/status")
async def update_document_status(doc_id: str, status: AgreementStatus, user: dict = Depends(get_current_user)):
    result = await db.generated_documents.update_one({"id": doc_id}, {"$set": {"status": status.value}})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Document not found")
    
    await log_activity(user['id'], user['name'], "update", "document", doc_id, f"Status changed to {status.value}")
    return {"message": "Document status updated"}

# ==================== DOCUMENT CHECKLIST ====================
@api_router.get("/checklist/{customer_id}")
async def get_checklist(customer_id: str, user: dict = Depends(get_current_user)):
    checklist = await db.document_checklists.find_one({"customer_id": customer_id}, {"_id": 0})
    if not checklist:
        # Create default checklist
        checklist = DocumentChecklist(customer_id=customer_id)
        doc = checklist.model_dump()
        doc['updated_at'] = doc['updated_at'].isoformat()
        await db.document_checklists.insert_one(doc)
        return doc
    return checklist

@api_router.put("/checklist/{customer_id}")
async def update_checklist(customer_id: str, items: Dict[str, bool], user: dict = Depends(get_current_user)):
    result = await db.document_checklists.update_one(
        {"customer_id": customer_id},
        {"$set": {"items": items, "updated_at": datetime.now(timezone.utc).isoformat()}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Checklist not found")
    
    await log_activity(user['id'], user['name'], "update", "checklist", customer_id, "Updated document checklist")
    return {"message": "Checklist updated"}

# ==================== COMMUNICATION ====================
@api_router.post("/communication/email")
async def send_email_notification(
    customer_id: str,
    subject: str,
    message: str,
    attachment_doc_id: Optional[str] = None,
    attachment_ids: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    recipient_email = customer.get('email')
    email_status = "pending"
    sendgrid_response = None
    attachments_info = []
    
    # Process attachments if provided
    attachment_list = []
    if attachment_ids:
        doc_ids = [id.strip() for id in attachment_ids.split(",") if id.strip()]
        for doc_id in doc_ids:
            # Try to find in generated_documents first
            doc = await db.generated_documents.find_one({"id": doc_id}, {"_id": 0})
            if doc:
                attachments_info.append(f"Generated: {doc.get('doc_type', 'document')}")
            else:
                # Try customer_documents
                doc = await db.customer_documents.find_one({"id": doc_id}, {"_id": 0})
                if doc:
                    attachments_info.append(f"Uploaded: {doc.get('filename', doc.get('doc_type', 'document'))}")
    
    # Build HTML email content with black and gold theme
    html_content = f"""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <style>
            @import url('https://fonts.googleapis.com/css2?family=Roboto:wght@300;400;500;700&display=swap');
        </style>
    </head>
    <body style="font-family: 'Roboto', Arial, sans-serif; line-height: 1.6; color: #1A1A1A; background: #f5f5f5; margin: 0; padding: 30px;">
        <div style="max-width: 650px; margin: 0 auto; background: #fff; border: 2px solid #D4AF37; border-radius: 8px; overflow: hidden;">
            <!-- Header -->
            <div style="background: #1A1A1A; padding: 20px; display: flex; align-items: center;">
                <div style="background: #D4AF37; color: #1A1A1A; padding: 10px 15px; border-radius: 6px; font-weight: bold; font-size: 18px; margin-right: 15px;">RRL</div>
                <div>
                    <div style="color: #D4AF37; font-size: 18px; font-weight: 700;">RRL Builders and Developers</div>
                    <div style="color: #999; font-size: 11px;">Beyond homes. A lifestyle</div>
                </div>
            </div>
            
            <!-- Content -->
            <div style="padding: 30px;">
                <p style="margin: 0 0 20px 0;">Dear {customer.get('name', 'Customer')},</p>
                <div style="white-space: pre-line; margin-bottom: 25px;">{message}</div>
                
                <!-- Signature -->
                <div style="margin-top: 30px; padding: 20px; background: #fafafa; border-radius: 8px;">
                    <div style="font-size: 15px; font-weight: 600; color: #1A1A1A; margin-bottom: 3px;">John</div>
                    <div style="font-size: 12px; color: #D4AF37; font-weight: 500; margin-bottom: 12px;">CRM MANAGER</div>
                    <div style="font-size: 12px; color: #666; line-height: 1.6;">
                        <strong>P:</strong> 9606579135<br>
                        <strong>E:</strong> <a href="mailto:crm@rrlbuildersanddevelopers.com" style="color: #D4AF37;">crm@rrlbuildersanddevelopers.com</a><br>
                        <strong>A:</strong> 4TH Floor, RRL Tower, Sompura gate, Sarjapura Bengaluru - 562125<br><br>
                        <a href="https://www.rrlbuildersanddevelopers.com" style="color: #D4AF37;">www.rrlbuildersanddevelopers.com</a>
                    </div>
                </div>
            </div>
            
            <!-- Footer -->
            <div style="background: #fafafa; padding: 15px; text-align: center; font-size: 11px; color: #888; border-top: 1px solid #e0e0e0;">
                <p style="margin: 0;">RRL Builders and Developers Pvt. Ltd. | <a href="https://www.rrlbuildersanddevelopers.com" style="color: #D4AF37;">www.rrlbuildersanddevelopers.com</a></p>
            </div>
        </div>
    </body>
    </html>
    """
    
    # Send via SendGrid if configured
    if SENDGRID_API_KEY:
        try:
            sg_message = Mail(
                from_email=(SENDGRID_FROM_EMAIL, SENDGRID_FROM_NAME),
                to_emails=recipient_email,
                subject=subject,
                html_content=html_content
            )
            
            sg = SendGridAPIClient(SENDGRID_API_KEY)
            response = sg.send(sg_message)
            
            if response.status_code in [200, 201, 202]:
                email_status = "sent"
                sendgrid_response = {"status_code": response.status_code}
                logger.info(f"Email sent to {recipient_email} - Subject: {subject}")
            else:
                email_status = "failed"
                sendgrid_response = {"status_code": response.status_code}
                logger.error(f"Failed to send email to {recipient_email} - Status: {response.status_code}")
                
        except Exception as e:
            email_status = "failed"
            sendgrid_response = {"error": str(e)}
            logger.error(f"SendGrid error: {str(e)}")
    else:
        email_status = "mocked (no API key)"
    
    # Build log content with attachment info
    log_content = f"To: {recipient_email}\nSubject: {subject}\n\n{message}"
    if attachments_info:
        log_content += f"\n\nAttachments:\n- " + "\n- ".join(attachments_info)
    
    # Log the communication
    log = CommunicationLog(
        customer_id=customer_id,
        channel="email",
        message_type=subject,
        content=log_content,
        status=email_status,
        sent_by=user['id']
    )
    
    doc = log.model_dump()
    doc['sent_at'] = doc['sent_at'].isoformat()
    await db.communication_logs.insert_one(doc)
    
    await log_activity(user['id'], user['name'], "send", "email", customer_id, f"Email: {subject} - Status: {email_status}")
    
    return {"message": f"Email {email_status}", "log_id": log.id, "email_status": email_status, "sendgrid_response": sendgrid_response}

@api_router.post("/communication/whatsapp")
async def send_whatsapp_notification(
    customer_id: str,
    message: str,
    user: dict = Depends(get_current_user)
):
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # MOCKED: In production, integrate with Twilio
    log = CommunicationLog(
        customer_id=customer_id,
        channel="whatsapp",
        message_type="notification",
        content=f"To: {customer['phone']}\n\n{message}",
        status="sent (MOCKED)",
        sent_by=user['id']
    )
    
    doc = log.model_dump()
    doc['sent_at'] = doc['sent_at'].isoformat()
    await db.communication_logs.insert_one(doc)
    
    await log_activity(user['id'], user['name'], "send", "whatsapp", customer_id, "WhatsApp notification")
    
    return {"message": "WhatsApp sent (MOCKED - Configure Twilio for production)", "log_id": log.id}

@api_router.get("/communication/{customer_id}")
async def get_communication_history(customer_id: str, user: dict = Depends(get_current_user)):
    logs = await db.communication_logs.find({"customer_id": customer_id}, {"_id": 0}).sort("sent_at", -1).to_list(100)
    return logs

@api_router.get("/email-logs")
async def get_all_email_logs(
    page: int = 1,
    limit: int = 50,
    status: Optional[str] = None,
    search: Optional[str] = None,
    user: dict = Depends(get_current_user)
):
    """Get all email logs across all customers for tracking"""
    query = {"channel": "email"}
    if status and status != "all":
        query["status"] = {"$regex": status, "$options": "i"}
    
    # Get total count
    total = await db.communication_logs.count_documents(query)
    
    # Fetch logs with pagination
    skip = (page - 1) * limit
    logs = await db.communication_logs.find(query, {"_id": 0}).sort("sent_at", -1).skip(skip).limit(limit).to_list(limit)
    
    # Enrich with customer names
    customer_ids = list(set(log.get("customer_id", "") for log in logs))
    customers = {}
    if customer_ids:
        customer_docs = await db.customers.find(
            {"id": {"$in": customer_ids}}, {"_id": 0, "id": 1, "name": 1, "email": 1, "customer_id": 1}
        ).to_list(len(customer_ids))
        customers = {c["id"]: c for c in customer_docs}
    
    enriched_logs = []
    for log in logs:
        cust = customers.get(log.get("customer_id", ""), {})
        log["customer_name"] = cust.get("name", "Unknown")
        log["customer_email"] = cust.get("email", "")
        log["customer_display_id"] = cust.get("customer_id", "")
        if search:
            search_lower = search.lower()
            if (search_lower not in log.get("customer_name", "").lower() and
                search_lower not in log.get("message_type", "").lower() and
                search_lower not in log.get("content", "").lower() and
                search_lower not in log.get("customer_email", "").lower()):
                continue
        enriched_logs.append(log)
    
    return {
        "logs": enriched_logs,
        "total": total,
        "page": page,
        "limit": limit,
        "total_pages": (total + limit - 1) // limit
    }

# ==================== GOOGLE FORMS WEBHOOK ====================
@api_router.post("/webhook/google-form")
async def google_form_webhook(data: GoogleFormWebhook, background_tasks: BackgroundTasks):
    # Create customer from form data
    customer = Customer(
        name=data.customer_name,
        phone=data.phone,
        email=data.email,
        project=data.project,
        tower=data.tower,
        unit_number=data.unit_number,
        father_name=data.father_name or "",
        pan_number=data.pan_number or "",
        booking_amount=data.booking_amount or 0,
        booking_date=data.booking_date or datetime.now().strftime("%Y-%m-%d"),
        agreement_status=AgreementStatus.DRAFT
    )
    customer.customer_id = await generate_customer_id()
    
    doc = customer.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await db.customers.insert_one(doc)
    
    # Create default checklist
    checklist = DocumentChecklist(customer_id=customer.id)
    checklist_doc = checklist.model_dump()
    checklist_doc['updated_at'] = checklist_doc['updated_at'].isoformat()
    await db.document_checklists.insert_one(checklist_doc)
    
    # Log welcome email (MOCKED)
    welcome_log = CommunicationLog(
        customer_id=customer.id,
        channel="email",
        message_type="welcome",
        content=f"Welcome to RRL Builders! Your booking for {data.project} - {data.unit_number} has been received.",
        status="sent (MOCKED)",
        sent_by="system"
    )
    log_doc = welcome_log.model_dump()
    log_doc['sent_at'] = log_doc['sent_at'].isoformat()
    await db.communication_logs.insert_one(log_doc)
    
    await log_activity("system", "System", "create", "customer", customer.id, f"Customer created via Google Form: {customer.name}")
    
    # Auto-generate booking transaction if booking_amount is set
    await auto_generate_booking_transaction(customer.id, doc, created_by="system")
    
    return {"message": "Customer created successfully", "customer_id": customer.customer_id}

# ==================== DASHBOARD ====================
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(user: dict = Depends(get_current_user)):
    total_customers = await db.customers.count_documents({})
    pending_agreements = await db.customers.count_documents({"agreement_status": {"$in": ["draft", "sent"]}})
    
    # Calculate payments
    today = datetime.now(timezone.utc).date()
    week_end = today + timedelta(days=7)
    
    # === REVENUE CALCULATION FROM INDIVIDUAL TRANSACTION RECORDS ===
    # Sum all transaction amounts from payment_transactions collection for dynamic updates
    transactions = await db.payment_transactions.find({}, {"_id": 0, "amount": 1}).to_list(100000)
    total_revenue = sum(t.get('amount', 0) or 0 for t in transactions)
    
    # Get total flat value from customers
    customers = await db.customers.find({}, {"_id": 0, "total_price": 1}).to_list(10000)
    total_flat_value = sum(c.get('total_price', 0) or 0 for c in customers)
    
    # Total revenue = sum of all transactions (booking amounts are already included as transactions)
    
    # Total balance = total flat value - total revenue collected
    total_balance = total_flat_value - total_revenue
    
    # Total pending = same as balance
    total_pending = total_balance
    
    # === PAYMENT SCHEDULE ANALYSIS (for due dates) ===
    schedules = await db.payment_schedules.find({}, {"_id": 0}).to_list(1000)
    
    payments_due_this_week = 0
    overdue_payments = 0
    payment_status_counts = {"pending": 0, "paid": 0, "overdue": 0, "partial": 0}
    
    for schedule in schedules:
        for item in schedule.get('items', []):
            status = item.get('payment_status', 'pending')
            payment_status_counts[status] = payment_status_counts.get(status, 0) + 1
            
            if status in ['pending', 'partial']:
                try:
                    due_date = datetime.strptime(item['due_date'], "%Y-%m-%d").date()
                    if due_date < today:
                        overdue_payments += 1
                    elif due_date <= week_end:
                        payments_due_this_week += 1
                except (ValueError, TypeError):
                    pass
    
    # Calculate pending percentage
    total_amount = total_revenue + total_pending
    pending_percentage = round((total_pending / total_amount * 100), 2) if total_amount > 0 else 0
    
    # Monthly revenue (last 6 months) - calculate from transaction dates
    monthly_revenue = []
    for i in range(5, -1, -1):
        month_date = datetime.now() - timedelta(days=30*i)
        month_name = month_date.strftime("%b")
        # For now, distribute evenly (can be enhanced with actual date-based calculation)
        monthly_revenue.append({"month": month_name, "revenue": total_revenue / 6 if total_revenue > 0 else 0})
    
    # === STAGE-BASED OVERDUE CALCULATION ===
    current_stage = None
    current_stage_name = None
    stage_overdue_count = 0
    stage_overdue_amount = 0
    overdue_customers_list = []
    
    settings = await db.settings.find_one({"type": "payment_stage"}, {"_id": 0})
    if settings and settings.get("current_stage"):
        stage_key = settings.get("current_stage")
        stage_info = next((s for s in PAYMENT_STAGES if s["key"] == stage_key), None)
        if stage_info:
            current_stage = stage_key
            current_stage_name = stage_info["name"]
            cumulative_percentage = stage_info["cumulative"]
            
            # Get all customers with full data for overdue calculation
            all_customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
            all_transactions = await db.payment_transactions.find({}, {"_id": 0}).to_list(100000)
            
            # Group transactions by customer
            txn_by_customer = {}
            for txn in all_transactions:
                cid = txn.get("customer_id")
                if cid not in txn_by_customer:
                    txn_by_customer[cid] = []
                txn_by_customer[cid].append(txn)
            
            for cust in all_customers:
                cust_id = cust.get("id")
                cust_total_price = cust.get("total_price", 0) or 0
                cust_booking_amount = cust.get("booking_amount", 0) or 0
                
                expected_amount = (cust_total_price * cumulative_percentage) / 100
                
                cust_txns = txn_by_customer.get(cust_id, [])
                txn_total = sum(t.get("amount", 0) or 0 for t in cust_txns)
                # Total received = transactions only (booking_amount is already recorded as transactions)
                total_received = txn_total
                
                overdue_amt = expected_amount - total_received
                if overdue_amt > 0:
                    stage_overdue_count += 1
                    stage_overdue_amount += overdue_amt
                    overdue_customers_list.append({
                        "customer_id": cust_id,
                        "customer_name": cust.get("name"),
                        "unit_number": cust.get("unit_number"),
                        "overdue_amount": round(overdue_amt, 2)
                    })
    
    return DashboardStats(
        total_customers=total_customers,
        pending_agreements=pending_agreements,
        payments_due_this_week=payments_due_this_week,
        overdue_payments=overdue_payments,
        total_revenue=total_revenue if user['role'] == 'admin' else 0,
        total_pending=total_pending if user['role'] == 'admin' else 0,
        total_flat_value=total_flat_value if user['role'] == 'admin' else 0,
        total_balance=total_balance if user['role'] == 'admin' else 0,
        pending_percentage=pending_percentage if user['role'] == 'admin' else 0,
        monthly_revenue=monthly_revenue if user['role'] == 'admin' else [],
        payment_status_breakdown=payment_status_counts,
        current_stage=current_stage if user['role'] == 'admin' else None,
        current_stage_name=current_stage_name if user['role'] == 'admin' else None,
        stage_overdue_count=stage_overdue_count if user['role'] == 'admin' else 0,
        stage_overdue_amount=round(stage_overdue_amount, 2) if user['role'] == 'admin' else 0,
        overdue_customers=overdue_customers_list[:10] if user['role'] == 'admin' else []
    )

@api_router.get("/dashboard/recent-activities")
async def get_recent_activities(limit: int = 20, user: dict = Depends(get_current_user)):
    activities = await db.activity_logs.find({}, {"_id": 0}).sort("timestamp", -1).limit(limit).to_list(limit)
    return activities

@api_router.get("/dashboard/upcoming-due-dates")
async def get_upcoming_due_dates(user: dict = Depends(get_current_user)):
    """
    Get customers with payment due dates in the next 5 days.
    Due date rule: 10 days from booking date.
    """
    today = datetime.now(timezone.utc).date()
    
    # Get all customers
    customers = await db.customers.find(
        {"stage": {"$ne": "pending_approval"}},
        {"_id": 0}
    ).to_list(1000)
    
    upcoming = []
    
    for customer in customers:
        booking_date_str = customer.get('booking_date')
        if not booking_date_str:
            continue
        
        try:
            # Parse booking date
            if isinstance(booking_date_str, str):
                booking_date = datetime.strptime(booking_date_str, "%Y-%m-%d").date()
            else:
                booking_date = booking_date_str
            
            # Due date is 10 days from booking date
            due_date = booking_date + timedelta(days=10)
            
            # Check if due date is within next 5 days (including overdue up to 3 days)
            days_until_due = (due_date - today).days
            
            if -3 <= days_until_due <= 5:
                upcoming.append({
                    "customer_id": customer.get('id'),
                    "customer_name": customer.get('name'),
                    "project": customer.get('project'),
                    "unit_number": customer.get('unit_number'),
                    "booking_date": booking_date.isoformat(),
                    "due_date": due_date.isoformat(),
                    "days_until_due": days_until_due,
                    "total_price": customer.get('total_price', 0),
                    "balance_amount": customer.get('balance_amount', 0),
                })
        except Exception as e:
            logger.error(f"Error parsing dates for customer {customer.get('id')}: {e}")
            continue
    
    # Sort by due date (closest first)
    upcoming.sort(key=lambda x: x['days_until_due'])
    
    return upcoming

# ==================== PAYMENT STAGE MANAGEMENT ====================
@api_router.get("/settings/payment-stages")
async def get_payment_stages(user: dict = Depends(get_current_user)):
    """Get available payment stages for admin selection"""
    return PAYMENT_STAGES

@api_router.get("/settings/current-stage")
async def get_current_stage(user: dict = Depends(get_current_user)):
    """Get the current payment stage set by admin"""
    settings = await db.settings.find_one({"type": "payment_stage"}, {"_id": 0})
    if not settings:
        return {"current_stage": None, "current_stage_name": None, "cumulative_percentage": 0}
    
    stage_key = settings.get("current_stage")
    stage_info = next((s for s in PAYMENT_STAGES if s["key"] == stage_key), None)
    
    return {
        "current_stage": stage_key,
        "current_stage_name": stage_info["name"] if stage_info else None,
        "cumulative_percentage": stage_info["cumulative"] if stage_info else 0,
        "updated_at": settings.get("updated_at"),
        "updated_by": settings.get("updated_by_name")
    }

@api_router.post("/settings/current-stage")
async def set_current_stage(data: dict, user: dict = Depends(check_role([UserRole.ADMIN]))):
    """Set the current payment stage (Admin only)"""
    stage_key = data.get("current_stage")
    
    if not stage_key:
        raise HTTPException(status_code=400, detail="current_stage is required")
    
    # Validate stage key
    stage_info = next((s for s in PAYMENT_STAGES if s["key"] == stage_key), None)
    if not stage_info:
        raise HTTPException(status_code=400, detail="Invalid stage key")
    
    await db.settings.update_one(
        {"type": "payment_stage"},
        {
            "$set": {
                "type": "payment_stage",
                "current_stage": stage_key,
                "updated_at": datetime.now(timezone.utc),
                "updated_by": user["id"],
                "updated_by_name": user.get("name", "Admin")
            }
        },
        upsert=True
    )
    
    await log_activity(user['id'], user['name'], "update", "settings", "payment_stage", f"Set current payment stage to: {stage_info['name']}")
    
    return {"message": "Payment stage updated", "current_stage": stage_key, "stage_name": stage_info["name"]}

@api_router.get("/dashboard/overdue-by-stage")
async def get_overdue_by_stage(user: dict = Depends(get_current_user)):
    """
    Get customers who are overdue based on the current payment stage.
    Overdue = Cumulative amount for current stage - Total payment received
    """
    # Get current stage setting
    settings = await db.settings.find_one({"type": "payment_stage"}, {"_id": 0})
    if not settings or not settings.get("current_stage"):
        return {
            "current_stage": None,
            "overdue_count": 0,
            "total_overdue_amount": 0,
            "overdue_customers": []
        }
    
    stage_key = settings.get("current_stage")
    stage_info = next((s for s in PAYMENT_STAGES if s["key"] == stage_key), None)
    if not stage_info:
        return {
            "current_stage": stage_key,
            "overdue_count": 0,
            "total_overdue_amount": 0,
            "overdue_customers": []
        }
    
    cumulative_percentage = stage_info["cumulative"]
    
    # Get all customers
    customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
    
    # Get all transactions
    all_transactions = await db.payment_transactions.find({}, {"_id": 0}).to_list(100000)
    txn_by_customer = {}
    for txn in all_transactions:
        cid = txn.get("customer_id")
        if cid not in txn_by_customer:
            txn_by_customer[cid] = []
        txn_by_customer[cid].append(txn)
    
    overdue_customers = []
    total_overdue_amount = 0
    
    for customer in customers:
        customer_id = customer.get("id")
        total_price = customer.get("total_price", 0) or 0
        booking_amount = customer.get("booking_amount", 0) or 0
        
        # Calculate expected amount for current stage
        expected_amount = (total_price * cumulative_percentage) / 100
        
        # Calculate total received
        customer_txns = txn_by_customer.get(customer_id, [])
        txn_total = sum(t.get("amount", 0) or 0 for t in customer_txns)
        # Total received = transactions only (booking_amount is already recorded as transactions)
        total_received = txn_total
        
        # Calculate overdue
        overdue_amount = expected_amount - total_received
        
        if overdue_amount > 0:
            overdue_customers.append({
                "customer_id": customer_id,
                "customer_name": customer.get("name"),
                "project": customer.get("project"),
                "unit_number": customer.get("unit_number"),
                "tower": customer.get("tower"),
                "total_price": total_price,
                "expected_amount": round(expected_amount, 2),
                "total_received": round(total_received, 2),
                "overdue_amount": round(overdue_amount, 2),
                "phone": customer.get("phone"),
                "email": customer.get("email"),
            })
            total_overdue_amount += overdue_amount
    
    # Sort by overdue amount (highest first)
    overdue_customers.sort(key=lambda x: x["overdue_amount"], reverse=True)
    
    return {
        "current_stage": stage_key,
        "current_stage_name": stage_info["name"],
        "cumulative_percentage": cumulative_percentage,
        "overdue_count": len(overdue_customers),
        "total_overdue_amount": round(total_overdue_amount, 2),
        "overdue_customers": overdue_customers
    }

# ==================== CUSTOMER NOTES ====================
@api_router.get("/customers/{customer_id}/notes")
async def get_customer_notes(customer_id: str, user: dict = Depends(get_current_user)):
    """Get all notes for a customer"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0, "notes": 1})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    return customer.get("notes", [])

@api_router.post("/customers/{customer_id}/notes")
async def add_customer_note(customer_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Add a note to a customer"""
    content = data.get("content", "").strip()
    if not content:
        raise HTTPException(status_code=400, detail="Note content is required")
    
    note = {
        "id": str(uuid.uuid4()),
        "content": content,
        "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["id"],
        "created_by_name": user.get("name", "Unknown")
    }
    
    result = await db.customers.update_one(
        {"id": customer_id},
        {"$push": {"notes": note}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    await log_activity(user['id'], user['name'], "create", "note", customer_id, f"Added note to customer")
    
    return note

@api_router.delete("/customers/{customer_id}/notes/{note_id}")
async def delete_customer_note(customer_id: str, note_id: str, user: dict = Depends(get_current_user)):
    """Delete a note from a customer"""
    result = await db.customers.update_one(
        {"id": customer_id},
        {"$pull": {"notes": {"id": note_id}}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Note not found")
    
    await log_activity(user['id'], user['name'], "delete", "note", customer_id, f"Deleted note from customer")
    
    return {"message": "Note deleted"}

# ==================== CUSTOMER PAYMENT DUE DATE ====================
@api_router.put("/customers/{customer_id}/payment-due-date")
async def update_payment_due_date(customer_id: str, data: dict, user: dict = Depends(get_current_user)):
    """Update payment due date for a customer"""
    due_date = data.get("payment_due_date")
    
    result = await db.customers.update_one(
        {"id": customer_id},
        {"$set": {"payment_due_date": due_date, "updated_at": datetime.now(timezone.utc)}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    await log_activity(user['id'], user['name'], "update", "payment_due_date", customer_id, f"Updated payment due date to {due_date}")
    
    return {"message": "Payment due date updated", "payment_due_date": due_date}

# ==================== CUSTOMER OVERDUE CALCULATION ====================
@api_router.get("/customers/{customer_id}/overdue")
async def get_customer_overdue(customer_id: str, user: dict = Depends(get_current_user)):
    """Get overdue amount for a specific customer based on current stage"""
    # Get current stage setting
    settings = await db.settings.find_one({"type": "payment_stage"}, {"_id": 0})
    if not settings or not settings.get("current_stage"):
        return {"overdue_amount": 0, "current_stage": None, "message": "No payment stage set by admin"}
    
    stage_key = settings.get("current_stage")
    stage_info = next((s for s in PAYMENT_STAGES if s["key"] == stage_key), None)
    if not stage_info:
        return {"overdue_amount": 0, "current_stage": stage_key}
    
    cumulative_percentage = stage_info["cumulative"]
    
    # Get customer
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    total_price = customer.get("total_price", 0) or 0
    booking_amount = customer.get("booking_amount", 0) or 0
    
    # Calculate expected amount for current stage
    expected_amount = (total_price * cumulative_percentage) / 100
    
    # Get transactions
    transactions = await db.payment_transactions.find({"customer_id": customer_id}, {"_id": 0}).to_list(1000)
    txn_total = sum(t.get("amount", 0) or 0 for t in transactions)
    # Total received = transactions only (booking_amount is already recorded as transactions)
    total_received = txn_total
    
    overdue_amount = max(0, expected_amount - total_received)
    
    return {
        "customer_id": customer_id,
        "current_stage": stage_key,
        "current_stage_name": stage_info["name"],
        "cumulative_percentage": cumulative_percentage,
        "expected_amount": round(expected_amount, 2),
        "total_received": round(total_received, 2),
        "overdue_amount": round(overdue_amount, 2),
        "is_overdue": overdue_amount > 0
    }

# ==================== OVERDUE CUSTOMERS LIST (for filtering) ====================
@api_router.get("/customers/overdue/list")
async def get_overdue_customers_list(user: dict = Depends(get_current_user)):
    """Get list of customer IDs who are overdue (for filtering)"""
    overdue_data = await get_overdue_by_stage(user)
    return {
        "customer_ids": [c["customer_id"] for c in overdue_data.get("overdue_customers", [])]
    }

# ==================== ACTIVITY LOGS ====================
@api_router.get("/activity-logs")
async def get_activity_logs(
    entity_type: Optional[str] = None,
    entity_id: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    user: dict = Depends(get_current_user)
):
    query = {}
    if entity_type:
        query["entity_type"] = entity_type
    if entity_id:
        query["entity_id"] = entity_id
    
    logs = await db.activity_logs.find(query, {"_id": 0}).sort("timestamp", -1).skip(skip).limit(limit).to_list(limit)
    return logs

# ==================== PROJECTS (for dropdowns) ====================
@api_router.get("/projects")
async def get_projects(user: dict = Depends(get_current_user)):
    # Return predefined RRL projects
    return [
        {"name": "RRL Palm Altezze", "location": "Varthur, Bangalore"},
        {"name": "RRL NC 216", "location": "Bangalore"},
        {"name": "RRL Palacio", "location": "Medahalli, Bangalore"},
        {"name": "RRL Nature Woods", "location": "Sarjapur, Bangalore"},
        {"name": "RRL Towers", "location": "Sarjapur"},
        {"name": "RRL Complex", "location": "Attibele Sarjapur Road"}
    ]

# ==================== UNIT PRICING ====================
@api_router.post("/units")
async def create_unit_pricing(unit: UnitPricingCreate, user: dict = Depends(check_role([UserRole.ADMIN, UserRole.MANAGER]))):
    """Create a new unit with pricing"""
    unit_doc = UnitPricing(
        **unit.model_dump(),
        uds=round(unit.saleable_area * 0.495046, 2)
    )
    doc = unit_doc.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    
    await db.units.insert_one(doc)
    await log_activity(user['id'], user['name'], "create", "unit", unit_doc.id, f"Created unit {unit.unit_number}")
    return {**doc, "_id": None}

@api_router.get("/units")
async def get_units(
    project: Optional[str] = None,
    tower: Optional[str] = None,
    bhk_type: Optional[str] = None,
    is_available: Optional[bool] = None,
    user: dict = Depends(get_current_user)
):
    """Get all units with optional filters"""
    query = {}
    if project:
        query["project"] = project
    if tower:
        query["tower"] = tower
    if bhk_type:
        query["bhk_type"] = bhk_type
    if is_available is not None:
        query["is_available"] = is_available
    
    units = await db.units.find(query, {"_id": 0}).to_list(1000)
    return units

@api_router.get("/units/{unit_id}")
async def get_unit(unit_id: str, user: dict = Depends(get_current_user)):
    unit = await db.units.find_one({"id": unit_id}, {"_id": 0})
    if not unit:
        raise HTTPException(status_code=404, detail="Unit not found")
    return unit

@api_router.put("/units/{unit_id}")
async def update_unit(unit_id: str, updates: Dict[str, Any], user: dict = Depends(check_role([UserRole.ADMIN, UserRole.MANAGER]))):
    # Recalculate UDS if saleable_area changed
    if 'saleable_area' in updates:
        updates['uds'] = round(updates['saleable_area'] * 0.495046, 2)
    
    result = await db.units.update_one({"id": unit_id}, {"$set": updates})
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Unit not found")
    
    await log_activity(user['id'], user['name'], "update", "unit", unit_id, "Updated unit")
    return {"message": "Unit updated"}

@api_router.post("/units/bulk-import")
async def bulk_import_units(units: List[UnitPricingCreate], user: dict = Depends(check_role([UserRole.ADMIN]))):
    """Bulk import units from Excel data"""
    created = 0
    for unit_data in units:
        unit_doc = UnitPricing(
            **unit_data.model_dump(),
            uds=round(unit_data.saleable_area * 0.495046, 2)
        )
        doc = unit_doc.model_dump()
        doc['created_at'] = doc['created_at'].isoformat()
        await db.units.insert_one(doc)
        created += 1
    
    await log_activity(user['id'], user['name'], "import", "units", "", f"Imported {created} units")
    return {"message": f"Imported {created} units successfully"}

# ==================== PUBLIC BOOKING FORM (No Auth Required) ====================
class BookingFormData(BaseModel):
    """Public booking form submission"""
    # Primary Applicant
    name: str
    phone: str
    email: EmailStr
    father_name: Optional[str] = None
    date_of_birth: Optional[str] = None
    gender: Optional[str] = None  # male, female, or spouse
    pan_number: Optional[str] = None
    aadhar_number: Optional[str] = None
    address: Optional[str] = None
    company: Optional[str] = None
    designation: Optional[str] = None
    profession: Optional[str] = None
    nationality: str = "Indian"
    
    # Co-Applicant (optional)
    co_applicant_name: Optional[str] = None
    co_applicant_father_name: Optional[str] = None
    co_applicant_phone: Optional[str] = None
    co_applicant_email: Optional[str] = None
    co_applicant_pan: Optional[str] = None
    co_applicant_aadhar: Optional[str] = None
    co_applicant_address: Optional[str] = None
    co_applicant_date_of_birth: Optional[str] = None
    co_applicant_profession: Optional[str] = None
    co_applicant_nationality: Optional[str] = "Indian"
    
    # Property Selection
    project: str
    tower: str
    unit_number: str
    bhk_type: Optional[str] = ""
    floor: int = 0
    saleable_area: float = 0
    rate_per_sqft: float = 0
    floor_rise_cost: float = 0  # Manual floor rise cost per sqft
    parking: Optional[str] = "1"
    additional_parking: int = 0
    
    # Calculated prices (from frontend)
    total_price: float = 0
    base_price: float = 0
    floor_rise_total: float = 0
    club_house_charges: float = 200000
    additional_parking_charges: float = 0
    labour_cess: float = 0
    gst_amount: float = 0
    
    # Initial Payment Info
    booking_amount: float = 0
    transaction_details: Optional[str] = None
    transaction_date: Optional[str] = None
    transaction_bank: Optional[str] = None
    
    # Finance Preference
    finance_type: str = "self"  # self, loan, mixed
    finance_bank: Optional[str] = None
    
    # Remarks
    remarks: Optional[str] = None

@api_router.post("/public/booking-form")
async def submit_booking_form(data: BookingFormData):
    """
    Public endpoint for booking form submission.
    Creates a customer with 'pending_approval' stage.
    """
    # Use data from frontend if provided, otherwise try to get from unit database
    unit = await db.units.find_one({
        "project": data.project,
        "tower": data.tower,
        "unit_number": data.unit_number
    }, {"_id": 0})
    
    # Use frontend data first, fallback to unit data if available
    rate_per_sqft = data.rate_per_sqft if data.rate_per_sqft > 0 else (unit.get('rate_per_sqft', 0) if unit else 0)
    saleable_area = data.saleable_area if data.saleable_area > 0 else (unit.get('saleable_area', 0) if unit else 0)
    floor = data.floor if data.floor > 0 else (unit.get('floor', 0) if unit else 0)
    bhk_type = data.bhk_type if data.bhk_type else (unit.get('bhk_type', '') if unit else '')
    uds = round(saleable_area * 0.495046, 2) if saleable_area > 0 else 0
    floor_rise_cost = data.floor_rise_cost if data.floor_rise_cost > 0 else 0
    
    # Use frontend calculated prices if available, otherwise calculate
    if data.total_price > 0:
        base_price = data.base_price
        floor_rise_total = data.floor_rise_total
        club_house = data.club_house_charges
        parking_charges = data.additional_parking_charges
        labour_cess = data.labour_cess
        gst = data.gst_amount
        total_price = data.total_price
    else:
        # Base price = Total Saleable Area × Rate/sqft
        base_price = rate_per_sqft * saleable_area
        # Floor rise is manual cost per sqft × saleable area
        floor_rise_total = floor_rise_cost * saleable_area
        club_house = 200000  # Default club house
        parking_charges = data.additional_parking * 300000  # ₹3,00,000 per additional parking
        subtotal = base_price + floor_rise_total + club_house + parking_charges
        labour_cess = subtotal * 0.007  # 0.70%
        gst = subtotal * 0.05  # 5%
        total_price = subtotal + labour_cess + gst
    
    customer = Customer(
        name=data.name,
        phone=data.phone,
        email=data.email,
        father_name=data.father_name or "",
        date_of_birth=data.date_of_birth,
        gender=data.gender or "male",
        pan_number=data.pan_number or "",
        aadhar_number=data.aadhar_number or "",
        address=data.address or "",
        company=data.company,
        designation=data.designation,
        nationality=data.nationality,
        co_applicant_name=data.co_applicant_name,
        co_applicant_father_name=data.co_applicant_father_name,
        co_applicant_phone=data.co_applicant_phone,
        co_applicant_email=data.co_applicant_email,
        co_applicant_pan=data.co_applicant_pan,
        co_applicant_aadhar=data.co_applicant_aadhar,
        co_applicant_address=data.co_applicant_address,
        project=data.project,
        tower=data.tower,
        unit_number=data.unit_number,
        floor=floor,
        bhk_type=bhk_type,
        saleable_area=saleable_area,
        uds=uds,
        parking=data.parking,
        additional_parking=data.additional_parking,
        rate_per_sqft=rate_per_sqft,
        base_price=round(base_price, 2),
        club_house_charges=round(club_house, 2),
        additional_parking_charges=round(parking_charges, 2),
        labour_cess=round(labour_cess, 2),
        gst_percentage=5,
        gst_amount=round(gst, 2),
        total_price=round(total_price, 2),
        booking_amount=data.booking_amount,
        booking_date=data.transaction_date or datetime.now().strftime("%Y-%m-%d"),
        total_received=data.booking_amount,
        balance_amount=round(total_price - data.booking_amount, 2),
        payment_received_percentage=round((data.booking_amount / total_price * 100) if total_price > 0 else 0, 2),
        payment_pending_percentage=round(100 - ((data.booking_amount / total_price * 100) if total_price > 0 else 0), 2),
        finance_type=data.finance_type,
        finance_bank=data.finance_bank,
        stage="pending_approval",  # Key: starts as pending
        agreement_status="draft",
        transaction_details=data.transaction_details,
        transaction_date=data.transaction_date,
        transaction_bank=data.transaction_bank,
        remarks=data.remarks,
        custom_fields={
            "profession": data.profession or "",
            "floor_rise_cost": floor_rise_cost,
            "floor_rise_total": round(floor_rise_total, 2),
            "co_applicant_profession": data.co_applicant_profession or "",
            "co_applicant_nationality": data.co_applicant_nationality or "Indian",
        }
    )
    customer.customer_id = await generate_customer_id()
    
    doc = customer.model_dump()
    doc['created_at'] = doc['created_at'].isoformat()
    doc['updated_at'] = doc['updated_at'].isoformat()
    
    await db.customers.insert_one(doc)
    
    # Create default checklist
    checklist = DocumentChecklist(customer_id=customer.id)
    checklist_doc = checklist.model_dump()
    checklist_doc['updated_at'] = checklist_doc['updated_at'].isoformat()
    await db.document_checklists.insert_one(checklist_doc)
    
    # Mark unit as unavailable
    if unit:
        await db.units.update_one(
            {"id": unit['id']},
            {"$set": {"is_available": False}}
        )
    
    await log_activity("system", "Booking Form", "create", "customer", customer.id, f"New booking: {customer.name} for {data.project} - {data.unit_number}")
    
    # === AUTO-SEND WELCOME EMAIL WITH PRICE BREAKUP ===
    email_status = "not_sent"
    if customer.email and SENDGRID_API_KEY:
        try:
            # Generate welcome email HTML
            welcome_html = generate_welcome_email_html(doc)
            
            # Generate price breakup HTML
            price_breakup_html = generate_price_breakup_html(doc)
            
            # Create email message
            subject = f"Welcome to {customer.project} - Booking Confirmation & Terms"
            message = Mail(
                from_email=(SENDGRID_FROM_EMAIL, SENDGRID_FROM_NAME),
                to_emails=customer.email,
                subject=subject,
                html_content=welcome_html
            )
            
            # Generate and attach Price Breakup PDF
            try:
                pdf_bytes = HTML(string=price_breakup_html).write_pdf()
                encoded_pdf = base64.b64encode(pdf_bytes).decode()
                
                filename = f"RRL_PriceBreakup_{customer.name.replace(' ', '_')}.pdf"
                attachment = Attachment(
                    FileContent(encoded_pdf),
                    FileName(filename),
                    FileType('application/pdf'),
                    Disposition('attachment')
                )
                message.add_attachment(attachment)
            except Exception as pdf_error:
                logger.error(f"Error generating PDF for auto-email: {str(pdf_error)}")
            
            # Send email
            sg = SendGridAPIClient(SENDGRID_API_KEY)
            response = sg.send(message)
            
            if response.status_code in [200, 201, 202]:
                email_status = "sent"
                logger.info(f"Auto-sent welcome email to {customer.email} for new booking")
                
                # Log communication
                log = CommunicationLog(
                    customer_id=customer.id,
                    channel="email",
                    message_type="Auto Welcome Email",
                    content=f"To: {customer.email}\nSubject: {subject}\n\n[Auto-sent on booking submission with Price Breakup PDF]",
                    status="sent",
                    sent_by="system"
                )
                log_doc = log.model_dump()
                log_doc['sent_at'] = log_doc['sent_at'].isoformat()
                await db.communication_logs.insert_one(log_doc)
            else:
                email_status = "failed"
                logger.error(f"Failed to auto-send welcome email: Status {response.status_code}")
                
        except Exception as e:
            email_status = "error"
            logger.error(f"Error auto-sending welcome email: {str(e)}")
    
    return {
        "message": "Booking submitted successfully! Our team will contact you shortly.",
        "customer_id": customer.customer_id,
        "reference_id": customer.id,
        "welcome_email_status": email_status
    }

# Public document upload endpoint (no auth required)
@api_router.post("/public/upload-document/{customer_id}")
async def public_upload_document(
    customer_id: str,
    doc_type: str = Form(...),
    file: UploadFile = File(...)
):
    """Public endpoint to upload documents for a customer during booking"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Read file content and convert to base64 for storage
    content = await file.read()
    base64_content = base64.b64encode(content).decode('utf-8')
    
    # Store in database
    doc_record = {
        "id": str(uuid.uuid4()),
        "customer_id": customer_id,
        "doc_type": doc_type,
        "filename": file.filename,
        "content_type": file.content_type,
        "content_base64": base64_content,
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "uploaded_by": "public_booking"
    }
    
    await db.customer_documents.insert_one(doc_record)
    
    # Update customer's uploaded_documents dict
    uploaded_docs = customer.get('uploaded_documents', {})
    uploaded_docs[doc_type] = doc_record['id']
    await db.customers.update_one(
        {"id": customer_id},
        {"$set": {"uploaded_documents": uploaded_docs}}
    )
    
    await log_activity("system", "Booking Form", "upload", "document", customer_id, f"Uploaded {doc_type}")
    return {"message": "Document uploaded", "doc_id": doc_record['id']}

# ==================== LEADS MANAGEMENT (Pending Approvals) ====================
@api_router.get("/leads/pending")
async def get_pending_leads(user: dict = Depends(get_current_user)):
    """Get all customers pending approval"""
    leads = await db.customers.find(
        {"stage": "pending_approval"},
        {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    return leads

@api_router.put("/leads/{customer_id}/approve")
async def approve_lead(customer_id: str, user: dict = Depends(check_role([UserRole.ADMIN, UserRole.MANAGER]))):
    """Approve a pending lead - moves to 'qualified' stage"""
    result = await db.customers.update_one(
        {"id": customer_id, "stage": "pending_approval"},
        {"$set": {
            "stage": "qualified",
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Lead not found or already approved")
    
    await log_activity(user['id'], user['name'], "approve", "lead", customer_id, "Lead approved and qualified")
    return {"message": "Lead approved successfully"}

@api_router.put("/leads/{customer_id}/reject")
async def reject_lead(customer_id: str, reason: str = "", user: dict = Depends(check_role([UserRole.ADMIN, UserRole.MANAGER]))):
    """Reject a pending lead"""
    # Get customer to release the unit
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Lead not found")
    
    # Release the unit
    await db.units.update_one(
        {"project": customer['project'], "tower": customer['tower'], "unit_number": customer['unit_number']},
        {"$set": {"is_available": True}}
    )
    
    # Delete the customer record
    await db.customers.delete_one({"id": customer_id})
    await db.document_checklists.delete_one({"customer_id": customer_id})
    
    await log_activity(user['id'], user['name'], "reject", "lead", customer_id, f"Lead rejected: {reason}")
    return {"message": "Lead rejected and removed"}

@api_router.put("/leads/{customer_id}/stage")
async def update_lead_stage(customer_id: str, stage: str, user: dict = Depends(get_current_user)):
    """Update customer stage"""
    valid_stages = ["pending_approval", "qualified", "agreement_pending", "agreement_done", "registration_done"]
    if stage not in valid_stages:
        raise HTTPException(status_code=400, detail=f"Invalid stage. Must be one of: {valid_stages}")
    
    result = await db.customers.update_one(
        {"id": customer_id},
        {"$set": {
            "stage": stage,
            "updated_at": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    await log_activity(user['id'], user['name'], "update", "customer_stage", customer_id, f"Stage changed to {stage}")
    return {"message": f"Stage updated to {stage}"}

# ==================== DOCUMENT UPLOAD ====================
@api_router.post("/customers/{customer_id}/upload-document")
async def upload_customer_document(
    customer_id: str,
    doc_type: str = Form(...),
    file: UploadFile = File(...),
    user: dict = Depends(get_current_user)
):
    """Upload a document for a customer (PAN, Aadhaar, etc.)"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    # Read file content and convert to base64 for storage
    content = await file.read()
    base64_content = base64.b64encode(content).decode('utf-8')
    
    # Store in database
    doc_record = {
        "id": str(uuid.uuid4()),
        "customer_id": customer_id,
        "doc_type": doc_type,
        "filename": file.filename,
        "content_type": file.content_type,
        "content_base64": base64_content,
        "uploaded_at": datetime.now(timezone.utc).isoformat(),
        "uploaded_by": user['id']
    }
    
    await db.customer_documents.insert_one(doc_record)
    
    # Update customer's uploaded_documents dict
    uploaded_docs = customer.get('uploaded_documents', {})
    uploaded_docs[doc_type] = doc_record['id']
    await db.customers.update_one(
        {"id": customer_id},
        {"$set": {"uploaded_documents": uploaded_docs}}
    )
    
    await log_activity(user['id'], user['name'], "upload", "document", customer_id, f"Uploaded {doc_type}")
    return {"message": "Document uploaded", "doc_id": doc_record['id']}

@api_router.get("/customers/{customer_id}/documents-list")
async def get_customer_uploaded_documents(customer_id: str, user: dict = Depends(get_current_user)):
    """Get list of uploaded documents for a customer"""
    docs = await db.customer_documents.find(
        {"customer_id": customer_id},
        {"_id": 0, "content_base64": 0}  # Exclude base64 for listing
    ).to_list(100)
    return docs

@api_router.get("/documents/download/{doc_id}")
async def download_document(doc_id: str, user: dict = Depends(get_current_user)):
    """Download a specific document"""
    doc = await db.customer_documents.find_one({"id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    # Decode base64 content
    content = base64.b64decode(doc['content_base64'])
    
    return Response(
        content=content,
        media_type=doc.get('content_type', 'application/octet-stream'),
        headers={
            "Content-Disposition": f"attachment; filename={doc['filename']}"
        }
    )

@api_router.get("/documents/preview/{doc_id}")
async def preview_document(doc_id: str, user: dict = Depends(get_current_user)):
    """Preview a document (returns base64 for frontend display)"""
    doc = await db.customer_documents.find_one({"id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    return {
        "id": doc['id'],
        "filename": doc['filename'],
        "content_type": doc.get('content_type', 'application/octet-stream'),
        "content_base64": doc['content_base64']
    }

# Delete generated document
@api_router.delete("/documents/{doc_id}")
async def delete_generated_document(doc_id: str, user: dict = Depends(get_current_user)):
    """Delete a generated document - restricted for accounts role"""
    # Accounts role cannot delete documents
    if user['role'] == 'accounts':
        raise HTTPException(status_code=403, detail="Accounts role cannot delete documents")
    
    doc = await db.generated_documents.find_one({"id": doc_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    await db.generated_documents.delete_one({"id": doc_id})
    await log_activity(user['id'], user['name'], "delete", "document", doc_id, f"Deleted generated document: {doc.get('doc_type')}")
    
    return {"message": "Document deleted successfully"}

# Delete uploaded customer document
@api_router.delete("/customers/{customer_id}/documents/{doc_id}")
async def delete_uploaded_document(customer_id: str, doc_id: str, user: dict = Depends(get_current_user)):
    """Delete an uploaded customer document - restricted for accounts role"""
    # Accounts role cannot delete documents
    if user['role'] == 'accounts':
        raise HTTPException(status_code=403, detail="Accounts role cannot delete documents")
    
    doc = await db.customer_documents.find_one({"id": doc_id, "customer_id": customer_id}, {"_id": 0})
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")
    
    await db.customer_documents.delete_one({"id": doc_id})
    
    # Update customer's uploaded_documents dict
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if customer:
        uploaded_docs = customer.get('uploaded_documents', {})
        # Remove the doc_id reference from uploaded_documents
        for key, val in list(uploaded_docs.items()):
            if val == doc_id:
                del uploaded_docs[key]
                break
        await db.customers.update_one(
            {"id": customer_id},
            {"$set": {"uploaded_documents": uploaded_docs}}
        )
    
    await log_activity(user['id'], user['name'], "delete", "uploaded_document", doc_id, f"Deleted uploaded document: {doc.get('filename', doc.get('doc_type'))}")
    
    return {"message": "Document deleted successfully"}

# ==================== HEALTH CHECK ====================
@api_router.get("/")
async def root():
    return {"message": "RRL Builders CRM API", "version": "1.0.0"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

# Root-level health endpoint for Kubernetes health probes
@app.get("/health")
async def root_health_check():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

# ==================== EXPORT FUNCTIONALITY ====================
@api_router.get("/export/customers/csv")
async def export_customers_csv(user: dict = Depends(check_role([UserRole.ADMIN, UserRole.MANAGER]))):
    """Export all customers data as CSV"""
    import io
    import csv
    
    customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
    
    if not customers:
        raise HTTPException(status_code=404, detail="No customers found")
    
    output = io.StringIO()
    
    # Define CSV headers
    headers = [
        "Customer ID", "Name", "Email", "Phone", "Project", "Tower", "Unit Number",
        "BHK Type", "Floor", "Saleable Area", "Rate/Sqft", "Base Price", "Floor Rise Cost",
        "Club House", "Additional Parking", "Labour Cess", "GST Amount", "Total Price",
        "Booking Amount", "Booking Date", "Total Received", "Balance Amount",
        "Payment Received %", "Agreement Status", "Stage", "Father Name", "PAN Number",
        "Address", "Created At"
    ]
    
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    
    for c in customers:
        writer.writerow({
            "Customer ID": c.get('customer_id', ''),
            "Name": c.get('name', ''),
            "Email": c.get('email', ''),
            "Phone": c.get('phone', ''),
            "Project": c.get('project', ''),
            "Tower": c.get('tower', ''),
            "Unit Number": c.get('unit_number', ''),
            "BHK Type": c.get('unit_type', ''),
            "Floor": c.get('floor', ''),
            "Saleable Area": c.get('saleable_area', 0),
            "Rate/Sqft": c.get('rate_per_sqft', 0),
            "Base Price": c.get('base_price', 0),
            "Floor Rise Cost": c.get('custom_fields', {}).get('floor_rise_cost', 0),
            "Club House": c.get('club_house_charges', 0),
            "Additional Parking": c.get('additional_parking_charges', 0),
            "Labour Cess": c.get('labour_cess', 0),
            "GST Amount": c.get('gst_amount', 0),
            "Total Price": c.get('total_price', 0),
            "Booking Amount": c.get('booking_amount', 0),
            "Booking Date": c.get('booking_date', ''),
            "Total Received": c.get('total_received', 0),
            "Balance Amount": c.get('balance_amount', 0),
            "Payment Received %": c.get('payment_received_percentage', 0),
            "Agreement Status": c.get('agreement_status', ''),
            "Stage": c.get('stage', ''),
            "Father Name": c.get('father_name', ''),
            "PAN Number": c.get('pan_number', ''),
            "Address": c.get('address', ''),
            "Created At": c.get('created_at', '')
        })
    
    csv_content = output.getvalue()
    output.close()
    
    await log_activity(user['id'], user['name'], "export", "customers", "all", "Exported customers to CSV")
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=RRL_Customers_Export.csv"}
    )

@api_router.get("/export/customers/excel")
async def export_customers_excel(user: dict = Depends(check_role([UserRole.ADMIN, UserRole.MANAGER]))):
    """Export all customers data as Excel"""
    import io
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export not available. Please install openpyxl.")
    
    customers = await db.customers.find({}, {"_id": 0}).to_list(10000)
    
    if not customers:
        raise HTTPException(status_code=404, detail="No customers found")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customers"
    
    # Define headers
    headers = [
        "Customer ID", "Name", "Email", "Phone", "Project", "Tower", "Unit Number",
        "BHK Type", "Floor", "Saleable Area", "Rate/Sqft", "Base Price", "Floor Rise Cost",
        "Club House", "Additional Parking", "Labour Cess", "GST Amount", "Total Price",
        "Booking Amount", "Booking Date", "Total Received", "Balance Amount",
        "Payment Received %", "Agreement Status", "Stage", "Created At"
    ]
    
    # Style headers with black and gold theme
    header_fill = PatternFill(start_color="1A1A1A", end_color="1A1A1A", fill_type="solid")
    header_font = Font(color="D4AF37", bold=True, name="Roboto")
    thin_border = Border(
        left=Side(style='thin', color='D4AF37'),
        right=Side(style='thin', color='D4AF37'),
        top=Side(style='thin', color='D4AF37'),
        bottom=Side(style='thin', color='D4AF37')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border
    
    # Add data
    for row, c in enumerate(customers, 2):
        data = [
            c.get('customer_id', ''),
            c.get('name', ''),
            c.get('email', ''),
            c.get('phone', ''),
            c.get('project', ''),
            c.get('tower', ''),
            c.get('unit_number', ''),
            c.get('unit_type', ''),
            c.get('floor', ''),
            c.get('saleable_area', 0),
            c.get('rate_per_sqft', 0),
            c.get('base_price', 0),
            c.get('custom_fields', {}).get('floor_rise_cost', 0),
            c.get('club_house_charges', 0),
            c.get('additional_parking_charges', 0),
            c.get('labour_cess', 0),
            c.get('gst_amount', 0),
            c.get('total_price', 0),
            c.get('booking_amount', 0),
            c.get('booking_date', ''),
            c.get('total_received', 0),
            c.get('balance_amount', 0),
            c.get('payment_received_percentage', 0),
            c.get('agreement_status', ''),
            c.get('stage', ''),
            c.get('created_at', '')
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.font = Font(name="Roboto")
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    await log_activity(user['id'], user['name'], "export", "customers", "all", "Exported customers to Excel")
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=RRL_Customers_Export.xlsx"}
    )

@api_router.get("/export/payments/csv")
async def export_payments_csv(user: dict = Depends(check_role([UserRole.ADMIN, UserRole.MANAGER, UserRole.ACCOUNTS]))):
    """Export all payment schedules as CSV"""
    import io
    import csv
    
    schedules = await db.payment_schedules.find({}, {"_id": 0}).to_list(10000)
    customers = {c['id']: c for c in await db.customers.find({}, {"_id": 0, "id": 1, "name": 1, "customer_id": 1, "project": 1, "unit_number": 1}).to_list(10000)}
    
    output = io.StringIO()
    headers = ["Customer ID", "Customer Name", "Project", "Unit", "Installment", "Milestone", "Amount", "Due Date", "Status", "Payment Date"]
    
    writer = csv.DictWriter(output, fieldnames=headers)
    writer.writeheader()
    
    for schedule in schedules:
        customer = customers.get(schedule.get('customer_id'), {})
        for item in schedule.get('items', []):
            writer.writerow({
                "Customer ID": customer.get('customer_id', ''),
                "Customer Name": customer.get('name', ''),
                "Project": customer.get('project', ''),
                "Unit": customer.get('unit_number', ''),
                "Installment": item.get('installment_name', ''),
                "Milestone": item.get('milestone', ''),
                "Amount": item.get('amount', 0),
                "Due Date": item.get('due_date', ''),
                "Status": item.get('payment_status', ''),
                "Payment Date": item.get('payment_date', '')
            })
    
    csv_content = output.getvalue()
    output.close()
    
    await log_activity(user['id'], user['name'], "export", "payments", "all", "Exported payments to CSV")
    
    return Response(
        content=csv_content,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=RRL_Payments_Export.csv"}
    )

# ==================== PAYMENT SCHEDULE PDF ====================
@api_router.post("/documents/generate-payment-schedule-pdf/{customer_id}")
async def generate_payment_schedule_pdf(customer_id: str, user: dict = Depends(get_current_user)):
    """Generate Payment Schedule PDF for a customer"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    schedule = await db.payment_schedules.find_one({"customer_id": customer_id}, {"_id": 0})
    schedule_items = schedule.get('items', []) if schedule else []
    
    if not schedule_items:
        raise HTTPException(status_code=404, detail="No payment schedule found. Please generate one first.")
    
    html_content = generate_payment_schedule_html(customer, schedule_items)
    
    await log_activity(user['id'], user['name'], "generate", "payment_schedule_pdf", customer_id, "Generated Payment Schedule PDF")
    
    return {
        "html": html_content,
        "filename": f"RRL_PaymentSchedule_{customer.get('name', 'Customer').replace(' ', '_')}.pdf"
    }


@api_router.post("/documents/generate-cost-breakup-pdf/{customer_id}")
async def generate_cost_breakup_pdf(customer_id: str, user: dict = Depends(get_current_user)):
    """Generate Cost Breakup PDF for a customer"""
    customer = await db.customers.find_one({"id": customer_id}, {"_id": 0})
    if not customer:
        raise HTTPException(status_code=404, detail="Customer not found")
    
    html_content = generate_cost_breakup_html(customer)
    
    await log_activity(user['id'], user['name'], "generate", "cost_breakup_pdf", customer_id, "Generated Cost Breakup PDF")
    
    return {
        "html": html_content,
        "filename": f"RRL_CostBreakup_{customer.get('name', 'Customer').replace(' ', '_')}.pdf"
    }


# ==================== ADMIN DATA RESET ENDPOINT ====================
@api_router.post("/admin/reset-data-with-seed")
async def reset_data_with_seed(user: dict = Depends(get_current_user)):
    """
    Admin-only endpoint to reset the database with seed data.
    This will DELETE all existing customers and transactions, then load seed data.
    """
    import json
    
    if user.get('role') != 'admin':
        raise HTTPException(status_code=403, detail="Admin access required")
    
    try:
        # Delete existing data
        customers_deleted = await db.customers.delete_many({})
        transactions_deleted = await db.payment_transactions.delete_many({})
        schedules_deleted = await db.payment_schedules.delete_many({})
        documents_deleted = await db.documents.delete_many({})
        
        logger.info(f"Deleted: {customers_deleted.deleted_count} customers, {transactions_deleted.deleted_count} transactions")
        
        # Load seed data
        seed_dir = os.path.dirname(os.path.abspath(__file__))
        
        customers_count = 0
        transactions_count = 0
        
        # Seed customers
        customers_file = os.path.join(seed_dir, "seed_data_customers.json")
        if os.path.exists(customers_file):
            with open(customers_file, "r") as f:
                customers_data = json.load(f)
            if customers_data:
                await db.customers.insert_many(customers_data)
                customers_count = len(customers_data)
                logger.info(f"Seeded {customers_count} customers into database")
        
        # Seed transactions
        transactions_file = os.path.join(seed_dir, "seed_data_transactions.json")
        if os.path.exists(transactions_file):
            with open(transactions_file, "r") as f:
                transactions_data = json.load(f)
            if transactions_data:
                await db.payment_transactions.insert_many(transactions_data)
                transactions_count = len(transactions_data)
                logger.info(f"Seeded {transactions_count} transactions into database")
        
        return {
            "success": True,
            "message": "Database reset and seeded successfully",
            "deleted": {
                "customers": customers_deleted.deleted_count,
                "transactions": transactions_deleted.deleted_count,
                "schedules": schedules_deleted.deleted_count,
                "documents": documents_deleted.deleted_count
            },
            "seeded": {
                "customers": customers_count,
                "transactions": transactions_count
            }
        }
    except Exception as e:
        logger.error(f"Error resetting data: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Error resetting data: {str(e)}")


# Include the router in the main app - MUST be after all route definitions
app.include_router(api_router)

# Include modular routers under /api prefix
# Phase 2: These routers will gradually replace the inline routes in api_router
app.include_router(auth_router, prefix="/api")
app.include_router(auth_admin_router, prefix="/api")
app.include_router(users_router, prefix="/api")
app.include_router(customers_router, prefix="/api")
app.include_router(schedule_router, prefix="/api")
app.include_router(transactions_router, prefix="/api")
app.include_router(calculator_router, prefix="/api")
app.include_router(dashboard_router, prefix="/api")

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def startup_event():
    import os
    import json
    
    # Create default admin user if not exists
    admin = await db.users.find_one({"email": "admin@rrlbuilders.com"})
    if not admin:
        admin_user = User(
            email="admin@rrlbuilders.com",
            name="Admin User",
            role=UserRole.ADMIN,
            phone="9876543210"
        )
        doc = admin_user.model_dump()
        doc['password_hash'] = hash_password("admin123")
        doc['created_at'] = doc['created_at'].isoformat()
        await db.users.insert_one(doc)
        logger.info("Default admin user created: admin@rrlbuilders.com / admin123")
    
    # Create RRL CRM Admin user if not exists
    crm_admin = await db.users.find_one({"email": "crm@rrlbuildersanddevelopers.com"})
    if not crm_admin:
        crm_admin_user = User(
            email="crm@rrlbuildersanddevelopers.com",
            name="RRL CRM Admin",
            role=UserRole.ADMIN,
            phone=None
        )
        doc = crm_admin_user.model_dump()
        doc['password_hash'] = hash_password("#RRLnew2026")
        doc['created_at'] = doc['created_at'].isoformat()
        await db.users.insert_one(doc)
        logger.info("RRL CRM Admin user created: crm@rrlbuildersanddevelopers.com")
    
    # Create Accounts role user if not exists
    accounts_user = await db.users.find_one({"email": "accounts@rrlbuilders.com"})
    if not accounts_user:
        accounts_user_obj = User(
            email="accounts@rrlbuilders.com",
            name="Accounts User",
            role=UserRole.ACCOUNTS,
            phone=None
        )
        doc = accounts_user_obj.model_dump()
        doc['password_hash'] = hash_password("accounts123")
        doc['created_at'] = doc['created_at'].isoformat()
        await db.users.insert_one(doc)
        logger.info("Accounts user created: accounts@rrlbuilders.com")
    
    # Seed customer data if database is empty
    customer_count = await db.customers.count_documents({})
    if customer_count == 0:
        logger.info("No customers found. Seeding customer data...")
        # Load seed data from files
        seed_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Seed customers
        customers_file = os.path.join(seed_dir, "seed_data_customers.json")
        if os.path.exists(customers_file):
            with open(customers_file, "r") as f:
                customers_data = json.load(f)
            if customers_data:
                await db.customers.insert_many(customers_data)
                logger.info(f"Seeded {len(customers_data)} customers into database")
        
        # Seed transactions
        transactions_file = os.path.join(seed_dir, "seed_data_transactions.json")
        if os.path.exists(transactions_file):
            with open(transactions_file, "r") as f:
                transactions_data = json.load(f)
            if transactions_data:
                await db.payment_transactions.insert_many(transactions_data)
                logger.info(f"Seeded {len(transactions_data)} transactions into database")
    else:
        logger.info(f"Database already has {customer_count} customers. Skipping seed.")
    
    # Create indexes
    await db.customers.create_index("customer_id", unique=True)
    await db.customers.create_index("email")
    await db.users.create_index("email", unique=True)
    
    # One-time migration: Auto-generate booking transactions for existing customers
    migration_flag = await db.settings.find_one({"type": "migration_booking_txn_done"})
    if not migration_flag:
        logger.info("Running one-time migration: auto-generate booking transactions...")
        all_customers = await db.customers.find(
            {"booking_amount": {"$gt": 0}}, {"_id": 0}
        ).to_list(10000)
        migrated = 0
        for cust in all_customers:
            cid = cust.get("id")
            ba = cust.get("booking_amount", 0) or 0
            if ba <= 0 or not cid:
                continue
            # Check if booking transactions already cover the amount (both legacy and new field names)
            existing = await db.payment_transactions.find(
                {"customer_id": cid, "$or": [
                    {"transaction_stage": "booking"},
                    {"transaction_type": "booking"}
                ]}, {"_id": 0, "amount": 1}
            ).to_list(1000)
            existing_sum = sum(t.get("amount", 0) or 0 for t in existing)
            if existing_sum >= ba:
                continue
            # Generate the missing booking transaction
            await auto_generate_booking_transaction(cid, cust, created_by="system-migration")
            migrated += 1
        logger.info(f"Migration complete: auto-generated booking transactions for {migrated} customers")
        await db.settings.insert_one({"type": "migration_booking_txn_done", "migrated_count": migrated, "run_at": datetime.now(timezone.utc).isoformat()})


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
